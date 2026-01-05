import pandas as pd
import os
import sys
import logging
from datetime import datetime, time as dtime
import time as tm
import re
import numpy as np
import unicodedata

# Importação do tkinter com tratamento de erro
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
    import sv_ttk
    import darkdetect
    TKINTER_AVAILABLE = True
except ImportError as e:
    TKINTER_AVAILABLE = False
    error_msg = """
╔════════════════════════════════════════════════════════════════╗
║  ERRO: tkinter não está disponível                            ║
╚════════════════════════════════════════════════════════════════╝

Para instalar o suporte ao tkinter no macOS com pyenv:

1. Instale o Tcl/Tk:
   brew install tcl-tk

2. Reinstale o Python com pyenv:
   env PATH="$(brew --prefix tcl-tk)/bin:$PATH" pyenv install --force 3.14.0
   
   (Substitua 3.14.0 pela sua versão do Python)

3. Ou execute o script de instalação automática:
   ./install_tkinter_macos.sh

4. Verifique a instalação:
   python -c "import tkinter; print('tkinter OK!')"

Erro detalhado: {0}
""".format(str(e))
    print(error_msg)
    logging.error(f"tkinter não está disponível: {str(e)}")
    sys.exit(1)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import traceback
from openpyxl import load_workbook
import shutil
# Imports para geração de PDF
from reportlab.lib.pagesizes import A4  # letter não utilizado, removido
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER  # TA_LEFT e TA_RIGHT não utilizados, removidos

# Configure logging to both file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('unified_processing.log'),
        logging.StreamHandler()
    ]
)

def check_excel_files_in_use(output_dir):
    """Verifica se há arquivos Excel em uso no diretório de saída"""
    try:
        excel_files = []
        for root, dirs, files in os.walk(output_dir):
            for file in files:
                if file.endswith('.xlsx'):
                    file_path = os.path.join(root, file)
                    try:
                        with open(file_path, 'r+b') as f:
                            pass
                    except PermissionError:
                        excel_files.append(file_path)
        
        if excel_files:
            logging.warning("Arquivos Excel em uso encontrados:")
            for file_path in excel_files:
                logging.warning(f"  - {file_path}")
            logging.warning("Por favor, feche estes arquivos no Excel antes de continuar.")
            return excel_files
        return []
    except Exception as e:
        logging.error(f"Erro ao verificar arquivos em uso: {e}")
        return []

def is_file_in_use(file_path):
    """Verifica se um arquivo específico está em uso"""
    try:
        if not os.path.exists(file_path):
            return False
        
        # Tentar abrir o arquivo em modo de escrita
        with open(file_path, 'r+b') as f:
            pass
        return False  # Arquivo não está em uso
    except PermissionError:
        return True  # Arquivo está em uso
    except Exception:
        return False  # Outros erros não indicam arquivo em uso

def get_alternative_filename(original_path, attempt=1):
    """Gera um nome alternativo para o arquivo se o original estiver em uso"""
    base_name = original_path.replace('.xlsx', '')
    timestamp = tm.strftime("%Y%m%d_%H%M%S")
    return f"{base_name}_{timestamp}_{attempt}.xlsx"

# --- Funções auxiliares otimizadas ---

def normalize_matricula(series):
    """Otimizado: normaliza matrícula de forma vetorizada"""
    return series.astype(str).str.strip().str.zfill(6)

# --- Classes dos scripts originais (adaptadas) ---

class CompanyProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.SUPPLY_FOLDER = os.path.join(base_dir, 'Integração_Abast')
        self.DRIVER_FOLDER = os.path.join(base_dir, 'Integração_Mot')
        self.OUTPUT_BASE_DIR = output_base_dir # Novo diretório base para saída
        self.version_suffix = version_suffix
        
    def find_available_companies(self):
        logging.info("Searching for available companies for Abst_Mot_Por_empresa...")
        
        if not os.path.exists(self.SUPPLY_FOLDER):
            logging.error(f"Supply folder not found: {self.SUPPLY_FOLDER}")
            return []
        if not os.path.exists(self.DRIVER_FOLDER):
            logging.error(f"Driver folder not found: {self.DRIVER_FOLDER}")
            return []

        supply_files = [f for f in os.listdir(self.SUPPLY_FOLDER) if f.startswith('Abastecimento_') and f.endswith('.xlsx')]
        driver_files = [f for f in os.listdir(self.DRIVER_FOLDER) if f.startswith('Motorista_') and f.endswith('.xlsx')]
        
        companies = set()
        
        for driver_file in driver_files:
            parts = driver_file.split('_')
            if len(parts) < 4:
                continue
                
            company = parts[1]
            month_year = '_'.join(parts[2:4]).replace('.xlsx', '')
            supply_file = f"Abastecimento_{company}_{month_year}.xlsx"
            
            if supply_file in supply_files:
                companies.add(company)
        
        return sorted(list(companies))
    
    def get_company_files(self, company):
        supply_files = [f for f in os.listdir(self.SUPPLY_FOLDER) if f.startswith(f'Abastecimento_{company}_') and f.endswith('.xlsx')]
        driver_files = [f for f in os.listdir(self.DRIVER_FOLDER) if f.startswith(f'Motorista_{company}_') and f.endswith('.xlsx')]
        
        pairs = []
        
        for driver_file in driver_files:
            parts = driver_file.split('_')
            if len(parts) < 4:
                continue
            month_year = '_'.join(parts[2:4]).replace('.xlsx', '')
            supply_file = f"Abastecimento_{company}_{month_year}.xlsx"
            
            if supply_file in supply_files:
                pairs.append({
                    'supply': os.path.join(self.SUPPLY_FOLDER, supply_file),
                    'drivers': os.path.join(self.DRIVER_FOLDER, driver_file),
                    'month_year': month_year
                })
        
        return pairs
    
    def normalize_column_name(self, col):
        return ''.join(c for c in unicodedata.normalize('NFD', str(col)) if unicodedata.category(c) != 'Mn').replace(' ', '').lower()
    
    def detect_time_format(self, series):
        valid_values = series.dropna()
        valid_values = valid_values[valid_values.astype(str).str.strip() != '']
        valid_values = valid_values[~valid_values.astype(str).str.strip().isin(['-----', '0m', 'nan', 'NaT'])]
        
        if len(valid_values) == 0:
            return None
        
        test_formats = [
            ' %H:%M:%S', ' %H:%M', ' %m-%d-%Y %H:%M:%S', ' %d/%m/%Y %H:%M:%S',
            ' %d/%m/%Y %H:%M', ' %Y-%m-%d %H:%M:%S', ' %Y-%m-%d %H:%M',
            ' %I:%M %p', ' %I:%M:%S %p'
        ]
        
        format_scores = {}
        sample = valid_values.head(min(10, len(valid_values)))
        
        for fmt in test_formats:
            successes = 0
            for value in sample:
                try:
                    datetime.strptime(str(value).strip(), fmt)
                    successes += 1
                except:
                    continue
            
            if successes > 0:
                format_scores[fmt] = successes / len(sample)
        
        if format_scores:
            best_format = max(format_scores.items(), key=lambda x: x[1])
            logging.info(f"Detected format: {best_format[0]} (success rate: {best_format[1]:.2%})")
            return best_format[0]
        
        logging.warning("Using regex detection as fallback.")
        return 'regex'
    
    def normalize_time_smart(self, value, detected_format=None, base_date=None):
        if pd.isnull(value) or value == '' or str(value).strip() in ['-----', '0m', 'nan', 'NaT']:
            return np.nan
        
        value_str = str(value).strip()
        
        if detected_format and detected_format != 'regex':
            try:
                dt_parsed = datetime.strptime(value_str, detected_format)
                
                if detected_format in [' %m-%d-%Y %H:%M:%S', ' %d/%m/%Y %H:%M:%S', ' %d/%m/%Y %H:%M', 
                                     ' %Y-%m-%d %H:%M:%S', ' %Y-%m-%d %H:%M']:
                    return dt_parsed
                
                elif base_date is not None:
                    if isinstance(base_date, str):
                        try:
                            base_date = datetime.strptime(base_date, '%d/%m/%Y').date()
                        except:
                            base_date = None
                    
                    if base_date:
                        return datetime.combine(base_date, dt_parsed.time())
                
                return dt_parsed.time()
                
            except Exception as e:
                logging.debug(f"Error using detected format {detected_format} for value '{value_str}': {e}")
        
        fallback_formats = [
            ' %H:%M:%S', ' %H:%M', ' %m-%d-%Y %H:%M:%S', ' %d/%m/%Y %H:%M:%S',
            ' %d/%m/%Y %H:%M', ' %Y-%m-%d %H:%M:%S', ' %Y-%m-%d %H:%M',
            ' %I:%M %p', ' %I:%M:%S %p', ' %H%M', ' %H%M%S'
        ]
        
        for fmt in fallback_formats:
            try:
                dt_parsed = datetime.strptime(value_str, fmt)
                
                if fmt in [' %m-%d-%Y %H:%M:%S', ' %d/%m/%Y %H:%M:%S', ' %d/%m/%Y %H:%M', 
                          ' %Y-%m-%d %H:%M:%S', ' %Y-%m-%d %H:%M']:
                    return dt_parsed
                
                elif base_date is not None:
                    if isinstance(base_date, str):
                        try:
                            base_date = datetime.strptime(base_date, '%d/%m/%Y').date()
                        except:
                            base_date = None
                    
                    if base_date:
                        return datetime.combine(base_date, dtime(dt_parsed.hour, dt_parsed.minute, dt_parsed.second))
                
                return dtime(dt_parsed.hour, dt_parsed.minute, dt_parsed.second)
                
            except:
                continue
        
        match = re.search(r'(\d{1,2}):(\d{2})(?::(\d{2}))?', value_str)
        if match:
            h, m, s = match.groups()
            try:
                h = int(h)
                m = int(m)
                s = int(s) if s else 0
                
                if base_date is not None:
                    if isinstance(base_date, str):
                        try:
                            base_date = datetime.strptime(base_date, '%d/%m/%Y').date()
                        except:
                            base_date = None
                    
                    if base_date:
                        return datetime.combine(base_date, dtime(h, m, s))
                
                return dtime(h, m, s)
            except:
                pass
        
        logging.warning(f"Could not normalize time: '{value_str}'")
        return np.nan
    
    def identify_time_columns(self, df):
        normalized_columns = {self.normalize_column_name(col): col for col in df.columns}
        
        start_names = ['pegada', 'inicio', 'start', 'horainicio', 'entrada', 'checkin']
        end_names = ['largada', 'fim', 'end', 'horafim', 'saida', 'checkout', 'final']
        
        start_col = None
        end_col = None
        
        for name in start_names:
            if name in normalized_columns:
                start_col = normalized_columns[name]
                break
        
        for name in end_names:
            if name in normalized_columns:
                end_col = normalized_columns[name]
                break
        
        logging.info(f"Identified columns - Start: {start_col}, End: {end_col}")
        return start_col, end_col
    
    def process_company_files(self, supply_file, driver_file, company, month_year):
        logging.info(f"\nProcessing company: {company} for {month_year}")
        start_time = tm.time()
        
        try:
            # Otimizado: usar engine explícito e otimizações de leitura
            df_supply = pd.read_excel(supply_file, engine='openpyxl')
            df_drivers = pd.read_excel(driver_file, engine='openpyxl')
            
            logging.debug(f"Supply file columns: {df_supply.columns.tolist()}")
            logging.debug(f"Driver file columns: {df_drivers.columns.tolist()}")
            
            start_col_original, end_col_original = self.identify_time_columns(df_drivers)
            
            if not start_col_original or not end_col_original:
                logging.error("ERROR: Could not identify start and/or end time columns!")
                logging.error(f"Available columns: {df_drivers.columns.tolist()}")
                return False
            
            logging.info(f"Detecting format for column '{start_col_original}'...")
            start_format = self.detect_time_format(df_drivers[start_col_original])
            
            logging.info(f"Detecting format for column '{end_col_original}'...")
            end_format = self.detect_time_format(df_drivers[end_col_original])
            
            df_supply.columns = [self.normalize_column_name(col) for col in df_supply.columns]
            df_drivers.columns = [self.normalize_column_name(col) for col in df_drivers.columns]
            
            if 'placa' in df_supply.columns:
                df_supply['placa'] = df_supply['placa'].astype(str)
            else:
                logging.warning("Column 'placa' not found in supply file. Skipping conversion.")
                df_supply['placa'] = '' 
            
            if 'placa' in df_drivers.columns:
                df_drivers['placa'] = df_drivers['placa'].astype(str)
            else:
                logging.warning("Column 'placa' not found in driver file. Skipping conversion.")
                df_drivers['placa'] = '' 
            
            for df in [df_supply, df_drivers]:
                for col in df.columns:
                    if col in ['matricula', 'matrícula', 'matricula.', 'matricula_']:
                        df.rename(columns={col: 'matricula'}, inplace=True)
            
            start_col_norm = self.normalize_column_name(start_col_original)
            end_col_norm = self.normalize_column_name(end_col_original)
            
            if start_col_norm in df_drivers.columns:
                df_drivers.rename(columns={start_col_norm: 'pegada'}, inplace=True)
            if end_col_norm in df_drivers.columns:
                df_drivers.rename(columns={end_col_norm: 'largada'}, inplace=True)
            
            if 'dia' in df_supply.columns:
                df_supply['Date'] = pd.to_datetime(df_supply['dia'], format='%d/%m/%Y', errors='coerce').dt.date
            else:
                logging.warning("Column 'dia' not found in supply file. Skipping date conversion.")
                df_supply['Date'] = pd.NaT 
            
            if 'dia' in df_drivers.columns:
                df_drivers['Date'] = pd.to_datetime(df_drivers['dia'], format='%d/%m/%Y', errors='coerce').dt.date
            else:
                logging.warning("Column 'dia' not found in driver file. Skipping date conversion.")
                df_drivers['Date'] = pd.NaT 
            
            logging.info("Normalizing start times...")
            df_drivers['pegada_dt'] = df_drivers.apply(
                lambda row: self.normalize_time_smart(row['pegada'], start_format, row['dia']) if 'pegada' in row and 'dia' in row else np.nan, 
                axis=1
            )
            
            logging.info("Normalizing end times...")
            df_drivers['largada_dt'] = df_drivers.apply(
                lambda row: self.normalize_time_smart(row['largada'], end_format, row['dia']) if 'largada' in row and 'dia' in row else np.nan, 
                axis=1
            )
            
            total_rows = len(df_drivers)
            valid_starts = df_drivers['pegada_dt'].notna().sum()
            valid_ends = df_drivers['largada_dt'].notna().sum()
            
            logging.info(f"\nConversion statistics:")
            logging.info(f"Total rows: {total_rows}")
            logging.info(f"Valid starts: {valid_starts} ({valid_starts/total_rows:.1%})" if total_rows > 0 else "Valid starts: 0 (0.0%)")
            logging.info(f"Valid ends: {valid_ends} ({valid_ends/total_rows:.1%})" if total_rows > 0 else "Valid ends: 0 (0.0%)")
            
            invalid_dates = df_drivers[df_drivers['Date'].isna()]
            if not invalid_dates.empty:
                logging.warning("\nATTENTION: Invalid dates found:")
                cols_to_display = [col for col in ['motorista', 'matricula', 'placa', 'dia'] if col in invalid_dates.columns]
                if cols_to_display:
                    logging.warning(invalid_dates[cols_to_display].to_string())
                else:
                    logging.warning("No relevant columns to display for invalid dates.")
            
            combinations = df_drivers[['placa', 'Date']].drop_duplicates()
            results = []
            total_combinations = len(combinations)
            logging.info(f"Processing {total_combinations} combinations...")

            for idx, row in combinations.iterrows():
                plate = row['placa']
                date = row['Date']

                driver_group = df_drivers[(df_drivers['placa'] == plate) & (df_drivers['Date'] == date)].copy()
                supply_group = df_supply[(df_supply['placa'] == plate) & (df_supply['Date'] == date)]

                try:
                    driver_group['pegada'] = driver_group['pegada_dt']
                    driver_group['largada'] = driver_group['largada_dt']
                    
                    # Otimizado: verificação mais eficiente de tipos datetime
                    valid_times = driver_group['pegada'].notna() & driver_group['largada'].notna()
                    # Verificar se são datetime apenas se necessário (mais eficiente que apply em todos)
                    if not valid_times.all():
                        valid_times = valid_times & (driver_group['pegada'].apply(lambda x: isinstance(x, datetime))) & \
                                      (driver_group['largada'].apply(lambda x: isinstance(x, datetime)))
                    
                    if not valid_times.all():
                        invalid_rows = driver_group[~valid_times]
                        if not invalid_rows.empty:
                            logging.debug(f"\n[LOG] Rows ignored due to invalid times for plate {plate} on {date}:")
                            cols_to_display = [col for col in ['motorista', 'matricula', 'placa', 'dia', 'pegada', 'largada'] if col in invalid_rows.columns]
                            if cols_to_display:
                                logging.debug(invalid_rows[cols_to_display].to_string())
                        driver_group = driver_group[valid_times]
                    
                    if driver_group.empty:
                        continue
                    
                    next_day_mask = driver_group['largada'] < driver_group['pegada']
                    driver_group.loc[next_day_mask, 'largada'] = \
                        driver_group.loc[next_day_mask, 'largada'] + pd.Timedelta(days=1)
                    
                    driver_group['duration'] = (driver_group['largada'] - driver_group['pegada']).dt.total_seconds() / 60
                    total_duration = driver_group['duration'].sum()
                    
                    invalid_durations = driver_group[driver_group['duration'] < 0]
                    if not invalid_durations.empty:
                        logging.warning(f"\nATTENTION: Negative durations found for plate {plate} on {date}:")
                        cols_to_display = [col for col in ['motorista', 'matricula', 'placa', 'Date', 'pegada', 'largada'] if col in invalid_durations.columns]
                        if cols_to_display:
                            logging.warning(invalid_durations[cols_to_display].to_string())
                        
                except Exception as e:
                    logging.error(f"\nError processing times for plate {plate} on {date}: {str(e)}")
                    continue
                
                total_km = supply_group['km'].sum() if 'km' in supply_group.columns else 0
                total_liters = supply_group['litros'].sum() if 'litros' in supply_group.columns else 0

                if not supply_group.empty and total_duration > 0:
                    # Distribuição proporcional baseada na duração
                    driver_group['km_distributed'] = (driver_group['duration'] / total_duration) * total_km
                    driver_group['liters_distributed'] = (driver_group['duration'] / total_duration) * total_liters
                    
                    # Ajuste final para garantir que a soma bata exatamente com os totais
                    km_distributed_sum = driver_group['km_distributed'].sum()
                    liters_distributed_sum = driver_group['liters_distributed'].sum()
                    
                    # Ajuste km_distributed
                    if abs(total_km - km_distributed_sum) > 0.01:
                        diff_km = total_km - km_distributed_sum
                        if km_distributed_sum > 0:
                            # Ajuste proporcional
                            proporcoes_km = driver_group['km_distributed'] / km_distributed_sum
                            ajuste_km = proporcoes_km * diff_km
                            driver_group['km_distributed'] += ajuste_km
                            driver_group['km_distributed'] = driver_group['km_distributed'].round(2)
                            
                            # Ajuste final para garantir soma exata
                            diff_final_km = total_km - driver_group['km_distributed'].sum()
                            if abs(diff_final_km) > 0.01:
                                idx_max = driver_group['km_distributed'].idxmax()
                                driver_group.at[idx_max, 'km_distributed'] += diff_final_km
                        else:
                            # Se não há distribuição, distribuir igualmente
                            driver_group['km_distributed'] = total_km / len(driver_group)
                    
                    # Ajuste liters_distributed
                    if abs(total_liters - liters_distributed_sum) > 0.01:
                        diff_liters = total_liters - liters_distributed_sum
                        if liters_distributed_sum > 0:
                            # Ajuste proporcional
                            proporcoes_liters = driver_group['liters_distributed'] / liters_distributed_sum
                            ajuste_liters = proporcoes_liters * diff_liters
                            driver_group['liters_distributed'] += ajuste_liters
                            driver_group['liters_distributed'] = driver_group['liters_distributed'].round(2)
                            
                            # Ajuste final para garantir soma exata
                            diff_final_liters = total_liters - driver_group['liters_distributed'].sum()
                            if abs(diff_final_liters) > 0.01:
                                idx_max = driver_group['liters_distributed'].idxmax()
                                driver_group.at[idx_max, 'liters_distributed'] += diff_final_liters
                        else:
                            # Se não há distribuição, distribuir igualmente
                            driver_group['liters_distributed'] = total_liters / len(driver_group)
                else:
                    driver_group['km_distributed'] = 0
                    driver_group['liters_distributed'] = 0
                
                results.append(driver_group)
            
            if not results:
                logging.warning(f"[LOG] No valid results for {company}!")
                return False
            
            df_final = pd.concat(results, ignore_index=True)
            
            for col in ['pegada', 'largada']:
                if col in df_final.columns:
                    df_final[col] = df_final[col].apply(
                        lambda x: x.strftime('%d/%m/%Y %H:%M') if isinstance(x, datetime) and pd.notna(x) else str(x) if pd.notna(x) else '')
            
            # Extrai mês e ano do month_year para uso no caminho de saída
            month, year = month_year.split('_')
            # Define o diretório de saída com base na empresa, ano e mês
            output_folder_path = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2))
            os.makedirs(output_folder_path, exist_ok=True)

            detailed_filename = f"Detalhado_{company}_{month_year}{self.version_suffix}.xlsx"
            consolidated_filename = f"Abst_Mot_Por_empresa_{company}_{month_year}{self.version_suffix}.xlsx"
            
            detailed_filepath = os.path.join(output_folder_path, detailed_filename)
            
            # Verificar e corrigir distribuição antes de salvar
            logging.info(f"🔍 Verificando qualidade da distribuição para {company} {month_year}...")
            self.verificar_e_corrigir_distribuicao(df_final, supply_file, detailed_filepath)
            
            # Otimizado: engine explícito para melhor performance
            df_final.to_excel(detailed_filepath, index=False, engine='openpyxl')
            
            self.create_consolidated_file(df_final, consolidated_filename, output_folder_path)
            
            total_time = tm.time() - start_time
            logging.info(f"Processing finished in {total_time:.2f} seconds")
            logging.info(f"Files generated:")
            logging.info(f"- {detailed_filepath}")
            logging.info(f"- {os.path.join(output_folder_path, consolidated_filename)}")
            
            return True
            
        except Exception as e:
            logging.error(f"Error processing files for {company}: {str(e)}")
            return False
    
    def verificar_e_corrigir_distribuicao(self, df_final, supply_file, detailed_filepath):
        """
        Verifica e corrige a distribuição dos valores km_distributed e liters_distributed
        para garantir que batam exatamente com os totais originais do arquivo de abastecimento
        """
        try:
            # Carregar dados de abastecimento originais
            df_supply = pd.read_excel(supply_file)
            
            # Calcular totais originais
            total_km_original = df_supply['km'].sum() if 'km' in df_supply.columns else 0
            total_litros_original = df_supply['litros'].sum() if 'litros' in df_supply.columns else 0
            
            # Calcular totais distribuídos
            total_km_distributed = df_final['km_distributed'].sum()
            total_litros_distributed = df_final['liters_distributed'].sum()
            
            # Calcular diferenças
            diff_km = total_km_original - total_km_distributed
            diff_litros = total_litros_original - total_litros_distributed
            
            # Verificar se há diferenças significativas
            tolerancia = 0.01  # 1 centésimo
            
            if abs(diff_km) > tolerancia or abs(diff_litros) > tolerancia:
                logging.info(f"🔧 Corrigindo distribuição - Diferença KM: {diff_km:.2f}, Litros: {diff_litros:.2f}")
                
                # Correção km_distributed
                if abs(diff_km) > tolerancia and total_km_distributed > 0:
                    proporcoes_km = df_final['km_distributed'] / total_km_distributed
                    ajuste_km = proporcoes_km * diff_km
                    df_final['km_distributed'] += ajuste_km
                    df_final['km_distributed'] = df_final['km_distributed'].round(2)
                    
                    # Ajuste final para garantir soma exata
                    diff_final_km = total_km_original - df_final['km_distributed'].sum()
                    if abs(diff_final_km) > tolerancia:
                        idx_max = df_final['km_distributed'].idxmax()
                        df_final.at[idx_max, 'km_distributed'] += diff_final_km
                    
                    logging.info(f"✅ KM corrigido - Diferença final: {diff_final_km:.2f}")
                
                # Correção liters_distributed
                if abs(diff_litros) > tolerancia and total_litros_distributed > 0:
                    proporcoes_litros = df_final['liters_distributed'] / total_litros_distributed
                    ajuste_litros = proporcoes_litros * diff_litros
                    df_final['liters_distributed'] += ajuste_litros
                    df_final['liters_distributed'] = df_final['liters_distributed'].round(2)
                    
                    # Ajuste final para garantir soma exata
                    diff_final_litros = total_litros_original - df_final['liters_distributed'].sum()
                    if abs(diff_final_litros) > tolerancia:
                        idx_max = df_final['liters_distributed'].idxmax()
                        df_final.at[idx_max, 'liters_distributed'] += diff_final_litros
                    
                    logging.info(f"✅ Litros corrigido - Diferença final: {diff_final_litros:.2f}")
                
                # Verificar se a correção funcionou
                total_km_final = df_final['km_distributed'].sum()
                total_litros_final = df_final['liters_distributed'].sum()
                
                logging.info(f"🎯 Distribuição corrigida com sucesso!")
                logging.info(f"   - KM: {total_km_final:.2f} (original: {total_km_original:.2f})")
                logging.info(f"   - Litros: {total_litros_final:.2f} (original: {total_litros_original:.2f})")
                
                # Verificação final de qualidade
                diff_final_km = abs(total_km_original - total_km_final)
                diff_final_litros = abs(total_litros_original - total_litros_final)
                
                if diff_final_km <= tolerancia and diff_final_litros <= tolerancia:
                    logging.info(f"✅ Verificação de qualidade: Distribuição corrigida com precisão!")
                else:
                    logging.warning(f"⚠️ Verificação de qualidade: Pequenas diferenças ainda presentes (KM: {diff_final_km:.4f}, Litros: {diff_final_litros:.4f})")
                    
            else:
                logging.info("✅ Distribuição já está correta, nenhuma correção necessária")
                
        except Exception as e:
            logging.error(f"❌ Erro ao verificar/corrigir distribuição: {str(e)}")
            logging.error(f"Traceback: {traceback.format_exc()}")

    def create_consolidated_file(self, df_detailed, filename, output_folder_path):
        if 'motorista' not in df_detailed.columns:
            df_detailed['motorista'] = 'Desconhecido'
        if 'matricula' not in df_detailed.columns:
            df_detailed['matricula'] = 'N/A'

        consolidated = df_detailed.groupby(['motorista', 'matricula']).agg(
            total_km=('km_distributed', 'sum'),
            total_liters=('liters_distributed', 'sum'),
            days_worked=('Date', 'nunique')
        ).reset_index()
        
        consolidated['avg_km_per_day'] = consolidated['total_km'] / consolidated['days_worked']
        consolidated = consolidated.sort_values('total_km', ascending=False)
        
        # Verificar qualidade dos dados consolidados
        self.verificar_qualidade_consolidado(consolidated, df_detailed)
        
        consolidated_filepath = os.path.join(output_folder_path, filename)
        # Otimizado: engine explícito
        consolidated.to_excel(consolidated_filepath, index=False, engine='openpyxl')
        
        return consolidated

    def verificar_qualidade_consolidado(self, consolidated, df_detailed):
        """
        Verifica a qualidade dos dados consolidados comparando com os dados detalhados
        """
        try:
            # Calcular totais dos dados detalhados
            total_km_detalhado = df_detailed['km_distributed'].sum()
            total_litros_detalhado = df_detailed['liters_distributed'].sum()
            
            # Calcular totais dos dados consolidados
            total_km_consolidado = consolidated['total_km'].sum()
            total_litros_consolidado = consolidated['total_liters'].sum()
            
            # Verificar se os totais batem
            diff_km = abs(total_km_detalhado - total_km_consolidado)
            diff_litros = abs(total_litros_detalhado - total_litros_consolidado)
            
            tolerancia = 0.01  # 1 centésimo
            
            if diff_km <= tolerancia and diff_litros <= tolerancia:
                logging.info(f"✅ Qualidade do consolidado: Totais batem perfeitamente")
                logging.info(f"   - KM detalhado: {total_km_detalhado:.2f}, consolidado: {total_km_consolidado:.2f}")
                logging.info(f"   - Litros detalhado: {total_litros_detalhado:.2f}, consolidado: {total_litros_consolidado:.2f}")
            else:
                logging.warning(f"⚠️ Qualidade do consolidado: Pequenas diferenças detectadas")
                logging.warning(f"   - KM: diferença de {diff_km:.4f}")
                logging.warning(f"   - Litros: diferença de {diff_litros:.4f}")
                
        except Exception as e:
            logging.error(f"❌ Erro ao verificar qualidade do consolidado: {str(e)}")

class RankingProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.RANKING_DIR = os.path.join(base_dir, "Ranking")
        self.TURNOS_DIR = os.path.join(base_dir, "Turnos_128")
        self.OUTPUT_BASE_DIR = output_base_dir # Novo diretório base para saída
        self.version_suffix = version_suffix
        
    def find_available_companies(self):
        logging.info("Searching for available companies for Ranking_Por_Empresa...")
        companies = set()
        
        if not os.path.exists(self.RANKING_DIR):
            logging.error(f"Ranking folder not found: {self.RANKING_DIR}")
            return []

        for ranking_file in os.listdir(self.RANKING_DIR):
            if ranking_file.startswith('Ranking_') and ranking_file.endswith('.xlsx'):
                parts = ranking_file.replace('.xlsx', '').split('_')
                if len(parts) >= 3:
                    company = parts[1] if parts[1] != 'Consolidado' else parts[2]
                    companies.add(company)
        
        return sorted(list(companies))
    
    def find_available_periods(self, company):
        periods = set()
        
        if not os.path.exists(self.RANKING_DIR):
            logging.error(f"Ranking folder not found: {self.RANKING_DIR}")
            return []

        for ranking_file in os.listdir(self.RANKING_DIR):
            if ranking_file.startswith(f'Ranking_{company}_') and ranking_file.endswith('.xlsx'):
                parts = ranking_file.replace('.xlsx', '').split('_')
                if len(parts) >= 4:
                    month = parts[-2]
                    year = parts[-1]
                    periods.add(f"{month}_{year}")
        
        return sorted(list(periods))
    
    def converter_formato_brasileiro(self, df, colunas):
        for coluna in colunas:
            if coluna in df.columns:
                df[coluna] = pd.to_numeric(df[coluna].astype(str).str.replace(' , ', '.'), errors='coerce')
        return df
    
    def process_company_period(self, company, month_year):
        try:
            month, year = month_year.split('_')
            
            # Diagnóstico: listar arquivos nas pastas de entrada
            logging.info(f"[DIAGNÓSTICO] Arquivos em {self.RANKING_DIR}: {os.listdir(self.RANKING_DIR) if os.path.exists(self.RANKING_DIR) else 'Pasta não encontrada'}")
            logging.info(f"[DIAGNÓSTICO] Arquivos em {self.TURNOS_DIR}: {os.listdir(self.TURNOS_DIR) if os.path.exists(self.TURNOS_DIR) else 'Pasta não encontrada'}")
            
            ranking_file = f"Ranking_{company}_{month}_{year}.xlsx"
            turnos_file = f"Turnos_128_{company}_{month}_{year}.xlsx"
            
            if not os.path.exists(os.path.join(self.RANKING_DIR, ranking_file)):
                raise FileNotFoundError(f"Arquivo de ranking não encontrado: {ranking_file}")
            
            if not os.path.exists(os.path.join(self.TURNOS_DIR, turnos_file)):
                raise FileNotFoundError(f"Arquivo de turnos não encontrado: {turnos_file}")
            
            logging.info(f"Processando {company} - {month_year}")
            
            # Otimizado: engine explícito para melhor performance
            df_ranking = pd.read_excel(os.path.join(self.RANKING_DIR, ranking_file), dtype=str, engine='openpyxl')
            df_turnos = pd.read_excel(os.path.join(self.TURNOS_DIR, turnos_file), dtype=str, engine='openpyxl')
            
            df_turnos = self.converter_formato_brasileiro(df_turnos, ['km'])
            if 'km/l' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['km/l'])
            if 'ponto acumulado' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['ponto acumulado'])
            if 'km' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['km'])

            # Otimizado: usar função auxiliar vetorizada
            df_ranking['matricula'] = normalize_matricula(df_ranking['matricula'])
            df_turnos['matricula'] = normalize_matricula(df_turnos['matricula'])

            turno_mais_rodou = df_turnos.groupby(['matricula', 'turno'])['km'].sum().reset_index()
            turno_mais_rodou = turno_mais_rodou.loc[turno_mais_rodou.groupby('matricula')['km'].idxmax()]
            turno_mais_rodou = turno_mais_rodou[['matricula', 'turno']]
            turno_mais_rodou.columns = ['matricula', 'Turno_Mais_Rodou']
            
            # Verifica se a coluna 'nm_linha' existe, senão usa 'linha'
            linha_col = 'nm_linha' if 'nm_linha' in df_turnos.columns else 'linha'
            linha_mais_rodou = df_turnos.groupby(['matricula', linha_col])['km'].sum().reset_index()
            linha_mais_rodou = linha_mais_rodou.loc[linha_mais_rodou.groupby('matricula')['km'].idxmax()]
            linha_mais_rodou = linha_mais_rodou[['matricula', linha_col, 'km']]
            linha_mais_rodou.columns = ['matricula', 'Linha_Mais_Rodou', 'KM_Linha']

            veiculo_mais_rodou = df_turnos.groupby(['matricula', 'placa'])['km'].sum().reset_index()
            veiculo_mais_rodou = veiculo_mais_rodou.loc[veiculo_mais_rodou.groupby('matricula')['km'].idxmax()]
            veiculo_mais_rodou = veiculo_mais_rodou[['matricula', 'placa', 'km']]
            veiculo_mais_rodou.columns = ['matricula', 'Veiculo_Mais_Rodou', 'KM_Veiculo']

            df_final = df_ranking.merge(turno_mais_rodou, on='matricula', how='left')
            df_final = df_final.merge(linha_mais_rodou, on='matricula', how='left')
            df_final = df_final.merge(veiculo_mais_rodou, on='matricula', how='left')
            
            if 'km' in df_final.columns and 'km/l' in df_final.columns:
                df_final['km'] = pd.to_numeric(df_final['km'], errors='coerce')
                df_final['km/l'] = pd.to_numeric(df_final['km/l'], errors='coerce')
                if 'Litros' in df_final.columns:
                    df_final = df_final.drop(columns=['Litros'])
                df_final['Litros'] = df_final['km'] / df_final['km/l']
                cols = list(df_final.columns)
                idx_kml = cols.index('km/l')
                insert_idx = idx_kml + 1
                cols.insert(insert_idx, cols.pop(cols.index('Litros')))
                df_final = df_final[cols]

            return df_final
        
        except Exception as e:
            logging.error(f"Erro ao processar {company} {month_year}: {str(e)}")
            return None
    
    def create_report(self, df_final, company, month_year):
        try:
            month, year = month_year.split('_')
            output_folder_path = os.path.join(self.OUTPUT_BASE_DIR, 'Ranking_Por_Empresa', company, year, month.zfill(2))
            os.makedirs(output_folder_path, exist_ok=True)
            output_file = os.path.join(output_folder_path, f'Ranking_Por_Empresa_{company}_{month}_{year}{self.version_suffix}.xlsx')
            # Caminho do arquivo Abst_Mot_Por_empresa
            abst_mot_file = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2), f"Abst_Mot_Por_empresa_{company}_{month}_{year}{self.version_suffix}.xlsx")
            # Caminho do arquivo Consolidado do Ranking_Km_Proporcional
            consolidado_km_prop_file = os.path.join(self.OUTPUT_BASE_DIR, 'Rankig_Km_Proporcional', company, year, month.zfill(2), f'Consolidado_{company}_{month}_{year}{self.version_suffix}.xlsx')
            # Carregar dados de Abst_Mot_Por_empresa se existir
            df_abst_mot = None
            if os.path.exists(abst_mot_file):
                df_abst_mot = pd.read_excel(abst_mot_file, engine='openpyxl')
                df_abst_mot['matricula'] = normalize_matricula(df_abst_mot['matricula'])
            # Carregar dados do consolidado do Ranking_Km_Proporcional se existir
            df_km_prop = None
            if os.path.exists(consolidado_km_prop_file):
                df_km_prop = pd.read_excel(consolidado_km_prop_file, engine='openpyxl')
                df_km_prop['matricula'] = normalize_matricula(df_km_prop['matricula'])
            # Função para adicionar e formatar colunas em cada aba
            def add_and_format_columns(df_sheet):
                df_sheet['matricula'] = normalize_matricula(df_sheet['matricula'])
                # Merge com Abst_Mot_Por_empresa
                if df_abst_mot is not None:
                    cols_to_merge = ['matricula']
                    if 'total_km' in df_abst_mot.columns:
                        cols_to_merge.append('total_km')
                    if 'total_liters' in df_abst_mot.columns:
                        cols_to_merge.append('total_liters')
                    if 'days_worked' in df_abst_mot.columns:
                        cols_to_merge.append('days_worked')
                    df_sheet = df_sheet.merge(df_abst_mot[cols_to_merge], on='matricula', how='left')
                    if 'total_km' in df_sheet.columns and 'total_liters' in df_sheet.columns:
                        df_sheet['Km/l_Int.'] = df_sheet['total_km'] / df_sheet['total_liters']
                    else:
                        df_sheet['Km/l_Int.'] = np.nan
                else:
                    df_sheet['total_km'] = np.nan
                    df_sheet['total_liters'] = np.nan
                    df_sheet['Km/l_Int.'] = np.nan
                    df_sheet['days_worked'] = np.nan
                # Merge com Consolidado do Ranking_Km_Proporcional
                if df_km_prop is not None:
                    cols_km_prop = ['matricula']
                    for col in ['km_distributed', 'liters_distributed', 'Km/l_Média']:
                        if col in df_km_prop.columns:
                            cols_km_prop.append(col)
                    df_sheet = df_sheet.merge(df_km_prop[cols_km_prop], on='matricula', how='left', suffixes=('', '_KmProp'))
                else:
                    df_sheet['km_distributed'] = np.nan
                    df_sheet['liters_distributed'] = np.nan
                    df_sheet['Km/l_Média'] = np.nan
                # Reorganizar colunas após KM_Veiculo
                colunas = list(df_sheet.columns)
                if 'KM_Veiculo' in colunas:
                    idx = colunas.index('KM_Veiculo') + 1
                    novas = ['total_km', 'total_liters', 'Km/l_Int.', 'days_worked', 'km_distributed', 'liters_distributed', 'Km/l_Média']
                    for n in novas:
                        if n in colunas:
                            colunas.remove(n)
                    for i, n in enumerate(novas):
                        if n in df_sheet.columns:
                            colunas.insert(idx + i, n)
                    df_sheet = df_sheet[colunas]
                return df_sheet
            
            # Verificar se o arquivo está em uso e tentar alternativas
            max_attempts = 10
            original_output_file = output_file
            
            for attempt in range(max_attempts):
                try:
                    # Verificar se o arquivo está em uso
                    if is_file_in_use(output_file):
                        if attempt < max_attempts - 1:
                            output_file = get_alternative_filename(original_output_file, attempt + 1)
                            logging.warning(f"Tentativa {attempt + 1}/{max_attempts}: Arquivo {original_output_file} está em uso. Tentando: {output_file}")
                            tm.sleep(1)
                            continue
                        else:
                            raise PermissionError(f"Arquivo {original_output_file} está em uso após {max_attempts} tentativas.")
                    
                    # Se chegou aqui, o arquivo não está em uso
                    break
                    
                except PermissionError as e:
                    if attempt < max_attempts - 1:
                        output_file = get_alternative_filename(original_output_file, attempt + 1)
                        logging.warning(f"Tentativa {attempt + 1}/{max_attempts}: Erro de permissão. Tentando: {output_file}")
                        tm.sleep(2)
                    else:
                        logging.error(f"Erro de permissão após {max_attempts} tentativas: {str(e)}")
                        raise Exception(f"Não foi possível criar o arquivo. Verifique se o arquivo está aberto no Excel: {original_output_file}")
            
            logging.info(f"Criando relatório: {output_file}")
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Todas as abas a serem criadas
                abas = {}
                # Aba principal
                df_ordenado = df_final.sort_values(
                    by=['Linha_Mais_Rodou', 'Turno_Mais_Rodou', 'km/l', 'motorista'],
                    ascending=[True, True, False, True]
                )
                abas['Todos'] = add_and_format_columns(df_ordenado)
                # Superior_Mediano
                abas['Superior_Mediano'] = add_and_format_columns(
                    df_ordenado[(df_ordenado['status'].isin(['Superior', 'Mediano'])) & (df_ordenado['km'] >= 1000)]
                )
                # Insuficiente
                abas['Insuficiente'] = add_and_format_columns(df_ordenado[df_ordenado['status'] == 'Insuficiente'])
                # Pontuacao_Baixa
                abas['Pontuacao_Baixa'] = add_and_format_columns(df_ordenado[df_ordenado['ponto acumulado'] <= 2])
                # Abas por linha
                for linha in df_ordenado['Linha_Mais_Rodou'].unique():
                    if pd.notna(linha):
                        df_linha = df_ordenado[
                            (df_ordenado['Linha_Mais_Rodou'] == linha) &
                            (df_ordenado['status'] != 'Insuficiente') &
                            (df_ordenado['km'] >= 1000) &
                            (df_ordenado['Turno_Mais_Rodou'].isin(['Manha', 'Manhã', 'Tarde']))
                        ].copy()
                        if not df_linha.empty:
                            sheet_name = f'Linha_{linha}'[:31]
                            df_manha = df_linha[df_linha['Turno_Mais_Rodou'].isin(['Manha', 'Manhã'])].copy()
                            df_tarde = df_linha[df_linha['Turno_Mais_Rodou'] == 'Tarde'].copy()
                            frames = []
                            if not df_manha.empty:
                                frames.append(df_manha)
                            if not df_manha.empty and not df_tarde.empty:
                                linha_branca = pd.DataFrame([[''] * len(df_linha.columns)], columns=df_linha.columns)
                                frames.append(linha_branca)
                                header_df = pd.DataFrame([df_linha.columns], columns=df_linha.columns)
                                frames.append(header_df)
                            if not df_tarde.empty:
                                frames.append(df_tarde)
                            if frames:
                                df_final_linha = pd.concat(frames, ignore_index=True)
                                abas[sheet_name] = add_and_format_columns(df_final_linha)
                # Escrever todas as abas
                for nome_aba, df_aba in abas.items():
                    df_aba.to_excel(writer, sheet_name=nome_aba, index=False, header=True)
                # Aplicar formatação condicional em todas as abas
                for nome_aba in abas.keys():
                    worksheet = writer.sheets[nome_aba]
                    header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                    def idx(col):
                        try:
                            return header.index(col)
                        except ValueError:
                            return None
                    idx_km = idx('km')
                    idx_dias = idx('dias')
                    idx_fase = idx('fase')
                    idx_status = idx('status')
                    idx_ponto = idx('ponto acumulado')
                    idx_km_veiculo = idx('KM_Veiculo')
                    idx_km_distributed = idx('km_distributed')
                    idx_total_km = idx('total_km')
                    idx_km = idx('km')
                    idx_days_worked = idx('days_worked')
                    idx_giro = idx('giro')
                    idx_freio = idx('freio')
                    idx_pedal = idx('pedal')
                    # Estilos
                    vermelho_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    branco_bold = Font(color='FFFFFF', bold=True)
                    amarelo_claro = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                    preto_negrito = Font(color='000000', bold=True)
                    verde_claro = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                    verde = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                    azul = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                    azul_escuro_negrito = Font(color='000080', bold=True)
                    amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    vermelho = Font(color='FF0000', bold=True)
                    # Medium border for header
                    medium_border = Border(left=Side(style='medium', color='000000'),
                                          right=Side(style='medium', color='000000'),
                                          top=Side(style='medium', color='000000'),
                                          bottom=Side(style='medium', color='000000'))
                    thin_border = Border(left=Side(style='thin', color='000000'),
                                        right=Side(style='thin', color='000000'),
                                        top=Side(style='thin', color='000000'),
                                        bottom=Side(style='thin', color='000000'))
                    # 2. Header formatting
                    header_targets = [
                        'matricula', 'motorista', 'km/l', 'Litros', 'giro', 'freio', 'pedal', 'fase', 'km', 'fechamento',
                        'ponto acumulado', 'status', 'empresa', 'dias', 'Turno_Mais_Rodou', 'Linha_Mais_Rodou', 'KM_Linha',
                        'Veiculo_Mais_Rodou', 'KM_Veiculo', 'total_km', 'total_liters', 'Km/l_Int.', 'days_worked',
                        'km_distributed', 'liters_distributed', 'Km/l_Média'
                    ]
                    for col_idx, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1))):
                        if header[col_idx] in header_targets:
                            cell.font = Font(bold=True)
                            cell.border = medium_border
                        else:
                            cell.border = thin_border
                    # 1, 3, 4, 5. Row formatting
                    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                        # 1. km, total_km, km_distributed < 900 (REMOVED KM_Veiculo)
                        for idx_col in [idx_km, idx_total_km, idx_km_distributed]:
                            if idx_col is not None and row[idx_col].value is not None:
                                try:
                                    val = float(str(row[idx_col].value).replace(',', '.'))
                                    if val < 900:
                                        cell = row[idx_col]
                                        cell.fill = vermelho_fill
                                        cell.font = branco_bold
                                except (ValueError, TypeError):
                                    pass
                        # 3. days_worked
                        if idx_days_worked is not None and row[idx_days_worked].value is not None:
                            try:
                                dias_val = int(float(str(row[idx_days_worked].value).replace(',', '.')))
                                cell = row[idx_days_worked]
                                if dias_val == 0:
                                    cell.fill = vermelho_fill
                                    cell.font = branco_bold
                                elif 1 <= dias_val <= 10:
                                    cell.fill = amarelo_claro
                                    cell.font = preto_negrito
                                elif 11 <= dias_val <= 15:
                                    cell.fill = verde_claro
                                    cell.font = preto_negrito
                                elif 16 <= dias_val <= 20:
                                    cell.fill = verde
                                    cell.font = preto_negrito
                                elif 21 <= dias_val <= 31:
                                    cell.fill = azul
                                    cell.font = azul_escuro_negrito
                            except (ValueError, TypeError):
                                pass
                        # 4. giro and freio >= 8
                        for idx_col in [idx_giro, idx_freio]:
                            if idx_col is not None and row[idx_col].value is not None:
                                try:
                                    val = float(str(row[idx_col].value).replace(',', '.'))
                                    if val >= 8:
                                        cell = row[idx_col]
                                        cell.font = vermelho
                                except (ValueError, TypeError):
                                    pass
                        # 5. pedal >= 16
                        if idx_pedal is not None and row[idx_pedal].value is not None:
                            try:
                                val = float(str(row[idx_pedal].value).replace(',', '.'))
                                if val >= 16:
                                    cell = row[idx_pedal]
                                    cell.font = vermelho
                            except (ValueError, TypeError):
                                pass
                        # ponto acumulado (condicional)
                        if idx_ponto is not None and idx_fase is not None and idx_status is not None and row[idx_ponto].value is not None:
                            try:
                                fase = row[idx_fase].value
                                status = row[idx_status].value
                                ponto = float(str(row[idx_ponto].value).replace(',', '.'))
                                if fase in ['Ouro', 'Ouro C'] and status == 'Mediano' and 3.97 <= ponto <= 3.99:
                                    cell = row[idx_ponto]
                                    cell.fill = amarelo
                                    cell.font = vermelho
                            except (ValueError, TypeError):
                                pass
                        # Bordas para todas as células da linha
                        for cell in row:
                            cell.border = thin_border
                        # dias column formatting (reapplied)
                        if idx_dias is not None and row[idx_dias].value is not None:
                            try:
                                dias_val = int(float(str(row[idx_dias].value).replace(',', '.')))
                                cell = row[idx_dias]
                                if dias_val == 0:
                                    cell.fill = vermelho_fill
                                    cell.font = branco_bold
                                elif 1 <= dias_val <= 10:
                                    cell.fill = amarelo_claro
                                    cell.font = preto_negrito
                                elif 11 <= dias_val <= 15:
                                    cell.fill = verde_claro
                                    cell.font = preto_negrito
                                elif 16 <= dias_val <= 20:
                                    cell.fill = verde
                                    cell.font = preto_negrito
                                elif 21 <= dias_val <= 31:
                                    cell.fill = azul
                                    cell.font = azul_escuro_negrito
                            except (ValueError, TypeError):
                                pass
            logging.info(f"Relatório gerado com sucesso: {output_file}")
            return output_file
        except PermissionError as e:
            error_msg = f"Erro de permissão ao criar relatório: {str(e)}"
            logging.error(error_msg)
            logging.error("Possíveis causas:")
            logging.error("1. O arquivo está aberto no Excel")
            logging.error("2. Permissões insuficientes na pasta de destino")
            logging.error("3. Antivírus bloqueando a criação do arquivo")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return None
        except FileNotFoundError as e:
            error_msg = f"Arquivo não encontrado: {str(e)}"
            logging.error(error_msg)
            logging.error("Verifique se os arquivos de entrada existem:")
            logging.error(f"- Ranking: {os.path.join(self.RANKING_DIR, f'Ranking_{company}_{month}_{year}.xlsx')}")
            logging.error(f"- Turnos: {os.path.join(self.TURNOS_DIR, f'Turnos_128_{company}_{month}_{year}.xlsx')}")
            return None
        except Exception as e:
            error_msg = f"Erro inesperado ao criar relatório: {str(e)}"
            logging.error(error_msg)
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return None

class RankingIntegracaoProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.RANKING_DIR = os.path.join(base_dir, "Ranking")
        self.TURNOS_DIR = os.path.join(base_dir, "Turnos_128")
        self.OUTPUT_BASE_DIR = output_base_dir # Novo diretório base para saída
        self.version_suffix = version_suffix

    def find_available_companies(self):
        logging.info("Searching for available companies for Ranking_Integração...")
        companies = set()
        
        # Busca empresas nos arquivos de Ranking
        if os.path.exists(self.RANKING_DIR):
            for ranking_file in os.listdir(self.RANKING_DIR):
                if ranking_file.startswith('Ranking_') and ranking_file.endswith('.xlsx'):
                    parts = ranking_file.replace('.xlsx', '').split('_')
                    if len(parts) >= 3:
                        company = parts[1] if parts[1] != 'Consolidado' else parts[2]
                        companies.add(company)
        
        return sorted(list(companies))

    def find_available_periods(self, company):
        periods = set()
        
        if not os.path.exists(self.RANKING_DIR):
            logging.error(f"Ranking folder not found: {self.RANKING_DIR}")
            return []

        for ranking_file in os.listdir(self.RANKING_DIR):
            if ranking_file.startswith(f'Ranking_{company}_') and ranking_file.endswith('.xlsx'):
                parts = ranking_file.replace('.xlsx', '').split('_')
                if len(parts) >= 4:
                    month = parts[-2]
                    year = parts[-1]
                    periods.add(f"{month}_{year}")
        
        return sorted(list(periods))

    def converter_formato_brasileiro(self, df, colunas):
        for coluna in colunas:
            if coluna in df.columns:
                df[coluna] = pd.to_numeric(df[coluna].astype(str).str.replace(' , ', '.'), errors='coerce')
        return df

    def encontrar_coluna_linha(self, df):
        for nome in ['linha', 'nm_linha', 'nome_linha', 'linha_nome']:
            if nome in df.columns:
                return nome
        return None

    def process_company_period(self, company, month_year):
        try:
            month, year = month_year.split('_')
            
            # Diagnóstico: listar arquivos nas pastas de entrada
            logging.info(f"[DIAGNÓSTICO] Arquivos em {self.RANKING_DIR}: {os.listdir(self.RANKING_DIR) if os.path.exists(self.RANKING_DIR) else 'Pasta não encontrada'}")
            logging.info(f"[DIAGNÓSTICO] Arquivos em {self.TURNOS_DIR}: {os.listdir(self.TURNOS_DIR) if os.path.exists(self.TURNOS_DIR) else 'Pasta não encontrada'}")
            
            ranking_file = f"Ranking_{company}_{month}_{year}.xlsx"
            turnos_file = f"Turnos_128_{company}_{month}_{year}.xlsx"
            abst_mot_file = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2), f"Abst_Mot_Por_empresa_{company}_{month}_{year}{self.version_suffix}.xlsx")
            detailed_file = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2), f"Detalhado_{company}_{month}_{year}{self.version_suffix}.xlsx")
            # Diagnóstico: listar arquivos nas pastas de saída
            output_abst_dir = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2))
            if os.path.exists(output_abst_dir):
                logging.info(f"[DIAGNÓSTICO] Arquivos em {output_abst_dir}: {os.listdir(output_abst_dir)}")
            else:
                logging.info(f"[DIAGNÓSTICO] Pasta de saída não encontrada: {output_abst_dir}")

            if not os.path.exists(os.path.join(self.RANKING_DIR, ranking_file)):
                raise FileNotFoundError(f"Arquivo de ranking não encontrado: {ranking_file}")
            
            if not os.path.exists(os.path.join(self.TURNOS_DIR, turnos_file)):
                raise FileNotFoundError(f"Arquivo de turnos não encontrado: {turnos_file}")

            # Carregar arquivos principais
            # Otimizado: engine explícito para melhor performance
            df_ranking = pd.read_excel(os.path.join(self.RANKING_DIR, ranking_file), dtype=str, engine='openpyxl')
            df_turnos = pd.read_excel(os.path.join(self.TURNOS_DIR, turnos_file), dtype=str, engine='openpyxl')

            # Padronizar campo matricula
            # Otimizado: usar função auxiliar vetorizada
            df_ranking['matricula'] = normalize_matricula(df_ranking['matricula'])
            df_turnos['matricula'] = normalize_matricula(df_turnos['matricula'])

            # Converter colunas numéricas
            df_turnos = self.converter_formato_brasileiro(df_turnos, ['km'])
            if 'km/l' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['km/l'])
            if 'ponto acumulado' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['ponto acumulado'])
            if 'km' in df_ranking.columns:
                df_ranking = self.converter_formato_brasileiro(df_ranking, ['km'])

            # Agrupamentos baseados no script de referência
            turno_mais_rodou = df_turnos.groupby(['matricula', 'turno'])['km'].sum().reset_index()
            turno_mais_rodou = turno_mais_rodou.loc[turno_mais_rodou.groupby('matricula')['km'].idxmax()]
            turno_mais_rodou = turno_mais_rodou[['matricula', 'turno']]
            turno_mais_rodou.columns = ['matricula', 'Turno_Mais_Rodou']

            col_linha_turnos = self.encontrar_coluna_linha(df_turnos)
            if not col_linha_turnos:
                raise Exception('Nenhuma coluna de linha encontrada em df_turnos!')
            linha_mais_rodou = df_turnos.groupby(['matricula', col_linha_turnos])['km'].sum().reset_index()
            linha_mais_rodou = linha_mais_rodou.loc[linha_mais_rodou.groupby('matricula')['km'].idxmax()]
            linha_mais_rodou = linha_mais_rodou[['matricula', col_linha_turnos, 'km']]
            linha_mais_rodou.columns = ['matricula', 'Linha_Mais_Rodou', 'KM_Linha']

            veiculo_mais_rodou = df_turnos.groupby(['matricula', 'placa'])['km'].sum().reset_index()
            veiculo_mais_rodou = veiculo_mais_rodou.loc[veiculo_mais_rodou.groupby('matricula')['km'].idxmax()]
            veiculo_mais_rodou = veiculo_mais_rodou[['matricula', 'placa', 'km']]
            veiculo_mais_rodou.columns = ['matricula', 'Veiculo_Mais_Rodou', 'KM_Veiculo']

            # Merge principal
            df_final = df_ranking.merge(turno_mais_rodou, on='matricula', how='left')
            df_final = df_final.merge(linha_mais_rodou, on='matricula', how='left')
            df_final = df_final.merge(veiculo_mais_rodou, on='matricula', how='left')

            # Adicionar coluna Litros após 'km/l' e 'Litros' (resultado de km/km_l)
            if 'km' in df_final.columns and 'km/l' in df_final.columns:
                df_final['km'] = pd.to_numeric(df_final['km'], errors='coerce')
                df_final['km/l'] = pd.to_numeric(df_final['km/l'], errors='coerce')
                if 'Litros' in df_final.columns:
                    df_final = df_final.drop(columns=['Litros'])
                df_final['Litros'] = df_final['km'] / df_final['km/l']
                cols = list(df_final.columns)
                idx_kml = cols.index('km/l')
                insert_idx = idx_kml + 1
                cols.insert(insert_idx, cols.pop(cols.index('Litros')))
                df_final = df_final[cols]

            # Adicionar informações de Abst_Mot_Por_empresa
            if os.path.exists(abst_mot_file):
                df_abst_mot = pd.read_excel(abst_mot_file, engine='openpyxl')
                df_abst_mot['matricula'] = normalize_matricula(df_abst_mot['matricula'])
                # Adicionar as colunas total_km, total_liters, days_worked
                cols_to_merge = ['matricula']
                if 'total_km' in df_abst_mot.columns:
                    cols_to_merge.append('total_km')
                if 'total_liters' in df_abst_mot.columns:
                    cols_to_merge.append('total_liters')
                if 'days_worked' in df_abst_mot.columns:
                    cols_to_merge.append('days_worked')
                df_final = df_final.merge(df_abst_mot[cols_to_merge], on='matricula', how='left')
                # Calcular Km/l_Int.
                if 'total_km' in df_final.columns and 'total_liters' in df_final.columns:
                    df_final['Km/l_Int.'] = df_final['total_km'] / df_final['total_liters']
                else:
                    df_final['Km/l_Int.'] = np.nan
            else:
                df_final['total_km'] = np.nan
                df_final['total_liters'] = np.nan
                df_final['Km/l_Int.'] = np.nan
                df_final['days_worked'] = np.nan

            # Reorganizar as colunas para inserir as novas após KM_Veiculo
            colunas = list(df_final.columns)
            if 'KM_Veiculo' in colunas:
                idx = colunas.index('KM_Veiculo') + 1
                novas = ['total_km', 'total_liters', 'Km/l_Int.', 'days_worked']
                for n in novas:
                    if n in colunas:
                        colunas.remove(n)
                for i, n in enumerate(novas):
                    if n in df_final.columns:
                        colunas.insert(idx + i, n)
                df_final = df_final[colunas]

            # Se o DataFrame final estiver vazio, criar uma linha em branco para evitar erro no Excel
            if df_final is None or df_final.empty:
                df_final = pd.DataFrame([{'matricula': '', 'motorista': '', 'km/l': '', 'Litros': '', 'giro': '', 'freio': '', 'pedal': '', 'fase': '', 'km': '', 'fechamento': '', 'ponto acumulado': '', 'status': '', 'dias': '', 'Turno_Mais_Rodou': '', 'Linha_Mais_Rodou': '', 'KM_Linha': '', 'Veiculo_Mais_Rodou': '', 'KM_Veiculo': '', 'total_km': '', 'total_liters': '', 'Km/l_Int.': '', 'days_worked': ''}])

            return df_final
        
        except Exception as e:
            logging.error(f"Erro ao processar Ranking_Integração para {company} {month_year}: {str(e)}")
            return None
    
    def create_report(self, df_final, company, month_year):
        try:
            month, year = month_year.split('_')
            output_folder_path = os.path.join(self.OUTPUT_BASE_DIR, 'Ranking_Integração', company, year, month.zfill(2))
            os.makedirs(output_folder_path, exist_ok=True)
            output_file = os.path.join(output_folder_path, f'Ranking_Integração_{company}_{month}_{year}{self.version_suffix}.xlsx')
            logging.info(f"Iniciando criação do relatório: {output_file}")
            # Verificar se o arquivo está em uso e tentar alternativas
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    if os.path.exists(output_file):
                        with open(output_file, 'r+b') as f:
                            pass
                    break
                except PermissionError:
                    if attempt < max_attempts - 1:
                        logging.warning(f"Tentativa {attempt + 1}/{max_attempts}: Arquivo {output_file} está em uso. Aguardando...")
                        tm.sleep(2)
                        base_name = output_file.replace('.xlsx', '')
                        output_file = f"{base_name}_{attempt + 1}.xlsx"
                    else:
                        logging.error(f"Arquivo {output_file} está em uso após {max_attempts} tentativas. Verifique se o arquivo está aberto no Excel.")
                        raise Exception(f"Arquivo em uso: {output_file}. Feche o arquivo no Excel e tente novamente.")
            logging.info(f"Criando relatório: {output_file}")
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df_ordenado = df_final.sort_values(
                    by=['Linha_Mais_Rodou', 'Turno_Mais_Rodou', 'km/l', 'motorista'],
                    ascending=[True, True, False, True]
                )
                df_ordenado.to_excel(writer, sheet_name='Todos', index=False)
                workbook = writer.book
                worksheet = writer.sheets['Todos']
                amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                vermelho = Font(color='FF0000')
                vermelho_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                branco_bold = Font(color='FFFFFF', bold=True)
                rosa_claro = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                vermelho_negrito = Font(color='FF0000', bold=True)
                amarelo_claro = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                preto_negrito = Font(color='000000', bold=True)
                azul_claro = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                azul_escuro_negrito = Font(color='000080', bold=True)
                verde_claro = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                verde_escuro_negrito = Font(color='006400', bold=True)
                header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                idx_total_km = header.index('total_km') if 'total_km' in header else None
                idx_days_worked = header.index('days_worked') if 'days_worked' in header else None
                idx_fase = header.index('fase') if 'fase' in header else None
                idx_status = header.index('status') if 'status' in header else None
                idx_ponto = header.index('ponto acumulado') if 'ponto acumulado' in header else None
                idx_km = header.index('km') if 'km' in header else None
                idx_dias = header.index('dias') if 'dias' in header else None
                formatted_count = 0
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    # Formatação para total_km < 900
                    if idx_total_km is not None and row[idx_total_km].value is not None:
                        try:
                            valor_total_km = float(str(row[idx_total_km].value).replace(',', '.'))
                            if valor_total_km < 900:
                                cell_km = row[idx_total_km]
                                cell_km.fill = vermelho_fill
                                cell_km.font = branco_bold
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    # Formatação para days_worked
                    if idx_days_worked is not None and row[idx_days_worked].value is not None:
                        try:
                            dias_val = int(float(str(row[idx_days_worked].value).replace(',', '.')))
                            cell_dias = row[idx_days_worked]
                            if dias_val == 0:
                                cell_dias.fill = rosa_claro
                                cell_dias.font = vermelho_negrito
                                formatted_count += 1
                            elif 1 <= dias_val <= 15:
                                cell_dias.fill = amarelo_claro
                                cell_dias.font = preto_negrito
                                formatted_count += 1
                            elif dias_val >= 20:
                                cell_dias.fill = verde_claro
                                cell_dias.font = verde_escuro_negrito
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    # Formatação original para dias
                    if idx_dias is not None and row[idx_dias].value is not None:
                        try:
                            dias_val = int(float(str(row[idx_dias].value).replace(',', '.')))
                            cell_dias = row[idx_dias]
                            if dias_val == 0:
                                cell_dias.fill = rosa_claro
                                cell_dias.font = vermelho_negrito
                                formatted_count += 1
                            elif 1 <= dias_val <= 10:
                                cell_dias.fill = amarelo_claro
                                cell_dias.font = preto_negrito
                                formatted_count += 1
                            elif 11 <= dias_val <= 19:
                                cell_dias.fill = azul_claro
                                cell_dias.font = azul_escuro_negrito
                                formatted_count += 1
                            elif dias_val >= 20:
                                cell_dias.fill = verde_claro
                                cell_dias.font = verde_escuro_negrito
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    # Formatação para fase Ouro/Ouro C com status Mediano e ponto entre 3.97-3.99
                    if (idx_ponto is not None and idx_fase is not None and idx_status is not None and row[idx_ponto].value is not None):
                        fase = row[idx_fase].value
                        status = row[idx_status].value
                        try:
                            ponto = float(str(row[idx_ponto].value).replace(',', '.'))
                            if fase in ['Ouro', 'Ouro C'] and status == 'Mediano' and 3.97 <= ponto <= 3.99:
                                cell = row[idx_ponto]
                                cell.fill = amarelo
                                cell.font = vermelho
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    # Formatação condicional para km < 900
                    if idx_km is not None and row[idx_km].value is not None:
                        try:
                            valor_km = float(str(row[idx_km].value).replace(',', '.'))
                            if valor_km < 900:
                                cell_km = row[idx_km]
                                cell_km.fill = vermelho_fill
                                cell_km.font = branco_bold
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                logging.info(f"Formatação aplicada em {formatted_count} células")
                logging.info(f"Relatório gerado com sucesso: {output_file}")
                return output_file
        except Exception as e:
            logging.error(f"Erro ao criar relatório: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return None

class RankingOuroMedianoProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.OUTPUT_BASE_DIR = output_base_dir
        self.version_suffix = version_suffix
        
    def find_available_companies(self):
        """Encontra empresas que têm relatórios Ranking_Por_Empresa gerados"""
        logging.info("Procurando empresas com relatórios Ranking_Por_Empresa para consolidação Ouro Mediano...")
        companies = set()
        
        ranking_por_empresa_dir = os.path.join(self.OUTPUT_BASE_DIR, 'Ranking_Por_Empresa')
        if not os.path.exists(ranking_por_empresa_dir):
            logging.warning(f"Diretório Ranking_Por_Empresa não encontrado: {ranking_por_empresa_dir}")
            return []
        
        for company_dir in os.listdir(ranking_por_empresa_dir):
            company_path = os.path.join(ranking_por_empresa_dir, company_dir)
            if os.path.isdir(company_path):
                # Verificar se há arquivos de relatório nesta empresa
                for year_dir in os.listdir(company_path):
                    year_path = os.path.join(company_path, year_dir)
                    if os.path.isdir(year_path):
                        for month_dir in os.listdir(year_path):
                            month_path = os.path.join(year_path, month_dir)
                            if os.path.isdir(month_path):
                                for file in os.listdir(month_path):
                                    if file.startswith('Ranking_Por_Empresa_') and file.endswith('.xlsx'):
                                        companies.add(company_dir)
                                        break
                            if company_dir in companies:
                                break
                    if company_dir in companies:
                        break
        
        return sorted(list(companies))
    
    def find_available_periods(self, company):
        """Encontra períodos disponíveis para uma empresa específica"""
        periods = set()
        
        ranking_por_empresa_dir = os.path.join(self.OUTPUT_BASE_DIR, 'Ranking_Por_Empresa', company)
        if not os.path.exists(ranking_por_empresa_dir):
            return []
        
        for year_dir in os.listdir(ranking_por_empresa_dir):
            year_path = os.path.join(ranking_por_empresa_dir, year_dir)
            if os.path.isdir(year_path):
                for month_dir in os.listdir(year_path):
                    month_path = os.path.join(year_path, month_dir)
                    if os.path.isdir(month_path):
                        for file in os.listdir(month_path):
                            if file.startswith('Ranking_Por_Empresa_') and file.endswith('.xlsx'):
                                # Extrair mês e ano do nome do arquivo
                                parts = file.replace('.xlsx', '').split('_')
                                if len(parts) >= 4:
                                    month = parts[-2]
                                    year = parts[-1]
                                    periods.add(f"{month}_{year}")
        
        return sorted(list(periods))
    
    def process_consolidation(self, selected_companies=None, selected_periods=None):
        """Processa a consolidação dos relatórios Ouro Mediano"""
        try:
            logging.info("Iniciando consolidação de relatórios Ouro Mediano...")
            
            all_data = []
            processed_files = 0
            
            # Se não especificadas, usar todas as empresas e períodos disponíveis
            if selected_companies is None:
                selected_companies = self.find_available_companies()
            
            for company in selected_companies:
                if selected_periods is None:
                    company_periods = self.find_available_periods(company)
                else:
                    company_periods = selected_periods
                
                for period in company_periods:
                    month, year = period.split('_')
                    file_path = os.path.join(
                        self.OUTPUT_BASE_DIR, 
                        'Ranking_Por_Empresa', 
                        company, 
                        year, 
                        month.zfill(2),
                        f'Ranking_Por_Empresa_{company}_{month}_{year}{self.version_suffix}.xlsx'
                    )
                    
                    if os.path.exists(file_path):
                        try:
                            logging.info(f"Processando arquivo: {file_path}")
                            
                            # Ler a aba 'Todos' do arquivo
                            df = pd.read_excel(file_path, sheet_name='Todos')
                            
                            # Adicionar colunas de identificação
                            df['Empresa'] = company
                            df['Periodo'] = period
                            
                            # Filtrar registros que atendem aos critérios
                            filtered_df = self.filter_ouro_mediano_records(df)
                            
                            if not filtered_df.empty:
                                all_data.append(filtered_df)
                                processed_files += 1
                                logging.info(f"Encontrados {len(filtered_df)} registros Ouro Mediano em {company} - {period}")
                            else:
                                logging.info(f"Nenhum registro Ouro Mediano encontrado em {company} - {period}")
                                
                        except Exception as e:
                            logging.error(f"Erro ao processar arquivo {file_path}: {str(e)}")
                            continue
                    else:
                        logging.warning(f"Arquivo não encontrado: {file_path}")
            
            if not all_data:
                logging.warning("Nenhum dado encontrado para consolidação")
                return None
            
            # Consolidar todos os dados em um único DataFrame
            consolidated_df = pd.concat(all_data, ignore_index=True)
            
            # Ordenar por empresa, período e km/l
            consolidated_df = consolidated_df.sort_values(
                by=['Empresa', 'Periodo', 'km/l', 'motorista'],
                ascending=[True, True, False, True]
            )
            
            logging.info(f"Consolidação concluída: {len(consolidated_df)} registros de {processed_files} arquivos")
            
            return consolidated_df
            
        except Exception as e:
            logging.error(f"Erro na consolidação: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return None
    
    def filter_ouro_mediano_records(self, df):
        """Filtra registros que atendem aos critérios Ouro Mediano"""
        try:
            # Verificar se as colunas necessárias existem
            required_columns = ['fase', 'status', 'ponto acumulado']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                logging.warning(f"Colunas ausentes no DataFrame: {missing_columns}")
                return pd.DataFrame()
            
            # Converter 'ponto acumulado' para numérico
            df['ponto_acumulado_numeric'] = pd.to_numeric(
                df['ponto acumulado'].astype(str).str.replace(',', '.'), 
                errors='coerce'
            )
            
            # Aplicar filtros
            filtered_df = df[
                (df['fase'].isin(['Ouro', 'Ouro C'])) &
                (df['status'] == 'Mediano') &
                (df['ponto_acumulado_numeric'] >= 3.97) &
                (df['ponto_acumulado_numeric'] <= 3.99)
            ].copy()
            
            # Remover coluna auxiliar
            if 'ponto_acumulado_numeric' in filtered_df.columns:
                filtered_df = filtered_df.drop(columns=['ponto_acumulado_numeric'])
            
            return filtered_df
            
        except Exception as e:
            logging.error(f"Erro ao filtrar registros Ouro Mediano: {str(e)}")
            return pd.DataFrame()
    
    def create_consolidated_report(self, df_consolidated, selected_periods=None, selected_companies=None):
        """Cria o relatório consolidado"""
        try:
            if df_consolidated is None or df_consolidated.empty:
                logging.warning("Nenhum dado para criar relatório consolidado")
                return False
            
            # Criar diretório de saída
            output_folder = os.path.join(self.OUTPUT_BASE_DIR, 'Ranking_Ouro_Mediano')
            os.makedirs(output_folder, exist_ok=True)
            
            # Gerar nome do arquivo baseado nos períodos e empresas
            if selected_periods:
                periods_sorted = sorted(selected_periods)
                first_period = periods_sorted[0]
                last_period = periods_sorted[-1]
                if selected_companies and len(selected_companies) == 1:
                    filename = f"Ranking_Ouro_Mediano_{selected_companies[0]}_{first_period}_a_{last_period}{self.version_suffix}.xlsx"
                else:
                    filename = f"Ranking_Ouro_Mediano_{first_period}_a_{last_period}{self.version_suffix}.xlsx"
            else:
                current_date = datetime.now().strftime("%Y%m%d")
                filename = f"Ranking_Ouro_Mediano_Consolidado_{current_date}{self.version_suffix}.xlsx"
            
            output_file = os.path.join(output_folder, filename)
            
            logging.info(f"Criando relatório consolidado: {output_file}")
            
            # Verificar se o arquivo está em uso
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    if os.path.exists(output_file):
                        with open(output_file, 'r+b') as f:
                            pass
                    break
                except PermissionError:
                    if attempt < max_attempts - 1:
                        logging.warning(f"Tentativa {attempt + 1}/{max_attempts}: Arquivo {output_file} está em uso. Aguardando...")
                        tm.sleep(2)
                        base_name = output_file.replace('.xlsx', '')
                        output_file = f"{base_name}_{attempt + 1}.xlsx"
                    else:
                        logging.error(f"Arquivo {output_file} está em uso após {max_attempts} tentativas.")
                        raise Exception(f"Arquivo em uso: {output_file}. Feche o arquivo no Excel e tente novamente.")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Salvar dados principais
                df_consolidated.to_excel(writer, sheet_name='Todos', index=False)
                
                # Aplicar formatação condicional
                workbook = writer.book
                worksheet = writer.sheets['Todos']
                
                # Definir estilos
                amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                vermelho = Font(color='FF0000')
                vermelho_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                branco_bold = Font(color='FFFFFF', bold=True)
                
                # Estilos para coluna dias
                rosa_claro = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                vermelho_negrito = Font(color='FF0000', bold=True)
                amarelo_claro = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
                preto_negrito = Font(color='000000', bold=True)
                azul_claro = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
                azul_escuro_negrito = Font(color='000080', bold=True)
                verde_claro = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                verde_escuro_negrito = Font(color='006400', bold=True)
                
                # Encontrar índices das colunas
                header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                idx_ponto = header.index('ponto acumulado') if 'ponto acumulado' in header else None
                idx_km = header.index('km') if 'km' in header else None
                idx_dias = header.index('dias') if 'dias' in header else None
                
                logging.info(f"Colunas encontradas: ponto={idx_ponto}, km={idx_km}, dias={idx_dias}")
                
                # Aplicar formatação condicional
                formatted_count = 0
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    # Formatação para ponto acumulado (todos já são 3.97-3.99)
                    if idx_ponto is not None and row[idx_ponto].value is not None:
                        try:
                            ponto = float(str(row[idx_ponto].value).replace(',', '.'))
                            if 3.97 <= ponto <= 3.99:
                                cell = row[idx_ponto]
                                cell.fill = amarelo
                                cell.font = vermelho
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    
                    # Formatação condicional para km < 900
                    if idx_km is not None and row[idx_km].value is not None:
                        try:
                            valor_km = float(str(row[idx_km].value).replace(',', '.'))
                            if valor_km < 900:
                                cell_km = row[idx_km]
                                cell_km.fill = vermelho_fill
                                cell_km.font = branco_bold
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                    
                    # Formatação para coluna dias
                    if idx_dias is not None and row[idx_dias].value is not None:
                        try:
                            dias_val = int(float(str(row[idx_dias].value).replace(',', '.')))
                            cell_dias = row[idx_dias]
                            
                            if dias_val == 0:
                                cell_dias.fill = rosa_claro
                                cell_dias.font = vermelho_negrito
                                formatted_count += 1
                            elif 1 <= dias_val <= 10:
                                cell_dias.fill = amarelo_claro
                                cell_dias.font = preto_negrito
                                formatted_count += 1
                            elif 11 <= dias_val <= 19:
                                cell_dias.fill = azul_claro
                                cell_dias.font = azul_escuro_negrito
                                formatted_count += 1
                            elif dias_val >= 20:
                                cell_dias.fill = verde_claro
                                cell_dias.font = verde_escuro_negrito
                                formatted_count += 1
                        except (ValueError, TypeError):
                            continue
                
                # Criar abas por empresa
                for empresa in df_consolidated['Empresa'].unique():
                    df_empresa = df_consolidated[df_consolidated['Empresa'] == empresa].copy()
                    sheet_name = f'Empresa_{empresa}'[:31]  # Limitar nome da aba
                    df_empresa.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Criar aba por período
                for periodo in df_consolidated['Periodo'].unique():
                    df_periodo = df_consolidated[df_consolidated['Periodo'] == periodo].copy()
                    sheet_name = f'Periodo_{periodo}'[:31]  # Limitar nome da aba
                    df_periodo.to_excel(writer, sheet_name=sheet_name, index=False)
                
                logging.info(f"Formatação aplicada em {formatted_count} células")
                logging.info(f"Relatório consolidado gerado com sucesso: {output_file}")
                return output_file
            
        except Exception as e:
            logging.error(f"Erro ao criar relatório consolidado: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return False

class RankingKmProporcionalProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.SUPPLY_FOLDER = os.path.join(base_dir, 'Integração_Abast')
        self.OUTPUT_BASE_DIR = output_base_dir
        self.version_suffix = version_suffix

    def find_available_companies(self):
        # Considera empresas a partir dos arquivos de abastecimento
        companies = set()
        if not os.path.exists(self.SUPPLY_FOLDER):
            logging.error(f"Pasta de abastecimento não encontrada: {self.SUPPLY_FOLDER}")
            return []
        for file in os.listdir(self.SUPPLY_FOLDER):
            if file.startswith('Abastecimento_') and file.endswith('.xlsx'):
                parts = file.replace('.xlsx', '').split('_')
                if len(parts) >= 3:
                    companies.add(parts[1])
        return sorted(list(companies))

    def find_available_periods(self, company):
        periods = set()
        if not os.path.exists(self.SUPPLY_FOLDER):
            return []
        for file in os.listdir(self.SUPPLY_FOLDER):
            if file.startswith(f'Abastecimento_{company}_') and file.endswith('.xlsx'):
                parts = file.replace('.xlsx', '').split('_')
                if len(parts) >= 4:
                    month = parts[2]
                    year = parts[3]
                    periods.add(f"{month}_{year}")
        return sorted(list(periods))

    def calcular_media_empresa(self, company, month_year):
        """Calcula o total de km, litros e km/l médio da empresa para o período."""
        supply_file = os.path.join(self.SUPPLY_FOLDER, f"Abastecimento_{company}_{month_year}.xlsx")
        if not os.path.exists(supply_file):
            logging.error(f"Arquivo de abastecimento não encontrado: {supply_file}")
            return None, None, None
        df = pd.read_excel(supply_file)
        total_km = df['km'].sum() if 'km' in df.columns else 0
        total_litros = df['litros'].sum() if 'litros' in df.columns else 0
        km_l_medio = total_km / total_litros if total_litros > 0 else 0
        return total_km, total_litros, km_l_medio

    def ajustar_km_distributed(self, detalhado_path, total_km):
        """Ajusta o km_distributed para que a soma bata exatamente com o total_km da empresa."""
        if not os.path.exists(detalhado_path):
            logging.error(f"Arquivo detalhado não encontrado: {detalhado_path}")
            return False
        df = pd.read_excel(detalhado_path)
        if 'km_distributed' not in df.columns:
            logging.error(f"Coluna 'km_distributed' não encontrada em {detalhado_path}")
            return False
        soma_atual = df['km_distributed'].sum()
        diff = total_km - soma_atual
        if abs(diff) < 1e-6:
            logging.info("Nenhum ajuste necessário em km_distributed.")
            return True
        # Ajuste proporcional
        if soma_atual == 0:
            logging.warning("Soma de km_distributed é zero, não é possível ajustar proporcionalmente.")
            return False
        proporcoes = df['km_distributed'] / soma_atual
        ajuste = proporcoes * diff
        df['km_distributed'] += ajuste
        # Corrigir possíveis arredondamentos
        df['km_distributed'] = df['km_distributed'].round(2)
        # Ajuste final para garantir soma exata
        diff_final = total_km - df['km_distributed'].sum()
        if abs(diff_final) > 0.01:
            # Corrige o maior valor
            idx_max = df['km_distributed'].idxmax()
            df.at[idx_max, 'km_distributed'] += diff_final
        # Salva o arquivo ajustado
        df.to_excel(detalhado_path, index=False, engine='openpyxl')
        logging.info(f"Ajuste proporcional realizado em {detalhado_path}. Diferença corrigida: {diff:.2f}")
        return True

    def ajustar_km_e_litros_distributed(self, detalhado_path, total_km, total_litros):
        """Ajusta km_distributed e liters_distributed para que as somas batam exatamente com os totais originais da empresa."""
        if not os.path.exists(detalhado_path):
            logging.error(f"Arquivo detalhado não encontrado: {detalhado_path}")
            return False
        df = pd.read_excel(detalhado_path)
        alterou = False
        # Ajuste km_distributed
        if 'km_distributed' in df.columns:
            soma_atual_km = df['km_distributed'].sum()
            diff_km = total_km - soma_atual_km
            if abs(diff_km) >= 1e-6 and soma_atual_km != 0:
                proporcoes_km = df['km_distributed'] / soma_atual_km
                ajuste_km = proporcoes_km * diff_km
                df['km_distributed'] += ajuste_km
                df['km_distributed'] = df['km_distributed'].round(2)
                diff_final_km = total_km - df['km_distributed'].sum()
                if abs(diff_final_km) > 0.01:
                    idx_max = df['km_distributed'].idxmax()
                    df.at[idx_max, 'km_distributed'] += diff_final_km
                alterou = True
                logging.info(f"Ajuste proporcional realizado em km_distributed. Diferença corrigida: {diff_km:.2f}")
        else:
            logging.warning(f"Coluna 'km_distributed' não encontrada em {detalhado_path}")
        # Ajuste liters_distributed
        if 'liters_distributed' in df.columns:
            soma_atual_litros = df['liters_distributed'].sum()
            diff_litros = total_litros - soma_atual_litros
            if abs(diff_litros) >= 1e-6 and soma_atual_litros != 0:
                proporcoes_litros = df['liters_distributed'] / soma_atual_litros
                ajuste_litros = proporcoes_litros * diff_litros
                df['liters_distributed'] += ajuste_litros
                df['liters_distributed'] = df['liters_distributed'].round(2)
                diff_final_litros = total_litros - df['liters_distributed'].sum()
                if abs(diff_final_litros) > 0.01:
                    idx_max = df['liters_distributed'].idxmax()
                    df.at[idx_max, 'liters_distributed'] += diff_final_litros
                alterou = True
                logging.info(f"Ajuste proporcional realizado em liters_distributed. Diferença corrigida: {diff_litros:.2f}")
        else:
            logging.warning(f"Coluna 'liters_distributed' não encontrada em {detalhado_path}")
        if alterou:
            df.to_excel(detalhado_path, index=False, engine='openpyxl')
        return alterou

    def process_company_period(self, company, month_year):
        # Calcula médias
        total_km, total_litros, km_l_medio = self.calcular_media_empresa(company, month_year)
        if total_km is None:
            return None
        year = month_year.split('_')[1]
        month = month_year.split('_')[0]
        # Caminho do detalhado (agora na pasta Rankig_Km_Proporcional)
        output_folder = os.path.join(self.OUTPUT_BASE_DIR, 'Rankig_Km_Proporcional', company, year, month.zfill(2))
        os.makedirs(output_folder, exist_ok=True)
        detalhado_path = os.path.join(
            output_folder,
            f'Detalhado_{company}_{month_year}{self.version_suffix}.xlsx'
        )
        # Copiar o detalhado original da pasta Abst_Mot_Por_empresa, se existir
        detalhado_origem = os.path.join(
            self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company, year, month.zfill(2),
            f'Detalhado_{company}_{month_year}{self.version_suffix}.xlsx'
        )
        if os.path.exists(detalhado_origem):
            import shutil
            shutil.copyfile(detalhado_origem, detalhado_path)
        else:
            logging.error(f"Arquivo detalhado de origem não encontrado: {detalhado_origem}")
            return None
        # Ajusta o km_distributed e liters_distributed no novo local
        self.ajustar_km_e_litros_distributed(detalhado_path, total_km, total_litros)
        # Gerar consolidado por motorista
        try:
            df = pd.read_excel(detalhado_path)
            if 'motorista' not in df.columns:
                df['motorista'] = 'Desconhecido'
            if 'matricula' not in df.columns:
                df['matricula'] = 'N/A'
            df['dias_trabalhados'] = 1
            agrupado = df.groupby(['motorista', 'matricula']).agg(
                km_distributed=('km_distributed', 'sum'),
                liters_distributed=('liters_distributed', 'sum'),
                dias_trabalhados=('dias_trabalhados', 'sum')
            ).reset_index()
            agrupado['Km/l_Média'] = agrupado['km_distributed'] / agrupado['liters_distributed']
            consolidado_path = os.path.join(output_folder, f'Consolidado_{company}_{month}_{year}{self.version_suffix}.xlsx')
            agrupado.to_excel(consolidado_path, index=False)
            logging.info(f"Consolidado por motorista salvo em: {consolidado_path}")
        except Exception as e:
            logging.error(f"Erro ao gerar consolidado por motorista: {e}")
        # Gera relatório consolidado (resumo)
        resumo_file = os.path.join(output_folder, f'Ranking_Km_Proporcional_{company}_{month}_{year}{self.version_suffix}.xlsx')
        resumo = pd.DataFrame({
            'Empresa': [company],
            'Período': [month_year],
            'Total_KM': [total_km],
            'Total_Litros': [total_litros],
            'KM/L_Médio': [km_l_medio]
        })
        resumo.to_excel(resumo_file, index=False)
        logging.info(f"Relatório Detalhado salvo em: {detalhado_path}")
        logging.info(f"Resumo Ranking_Km_Proporcional salvo em: {resumo_file}")
        return detalhado_path

class TurnosIntegracaoProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.OUTPUT_BASE_DIR = output_base_dir
        self.version_suffix = version_suffix
        
        # Definição dos turnos conforme especificado
        self.turnos_definicao = {
            'Madrugada': {'inicio': 0, 'fim': 5, 'hora_inicio': '00:00', 'hora_fim': '05:59'},
            'Manhã': {'inicio': 6, 'fim': 11, 'hora_inicio': '06:00', 'hora_fim': '11:59'},
            'Intervalo': {'inicio': 12, 'fim': 13, 'hora_inicio': '12:00', 'hora_fim': '13:59'},
            'Tarde': {'inicio': 14, 'fim': 19, 'hora_inicio': '14:00', 'hora_fim': '19:59'},
            'Noite': {'inicio': 20, 'fim': 23, 'hora_inicio': '20:00', 'hora_fim': '23:59'}
        }
    
    def find_available_companies(self):
        """Encontra empresas que têm arquivos Detalhado disponíveis"""
        logging.info("Procurando empresas com arquivos Detalhado para processamento de Turnos Integração...")
        companies = set()
        
        abst_mot_dir = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa')
        if not os.path.exists(abst_mot_dir):
            logging.warning(f"Diretório Abst_Mot_Por_empresa não encontrado: {abst_mot_dir}")
            return []
        
        for company_dir in os.listdir(abst_mot_dir):
            company_path = os.path.join(abst_mot_dir, company_dir)
            if os.path.isdir(company_path):
                # Verificar se há arquivos Detalhado nesta empresa
                for year_dir in os.listdir(company_path):
                    year_path = os.path.join(company_path, year_dir)
                    if os.path.isdir(year_path):
                        for month_dir in os.listdir(year_path):
                            month_path = os.path.join(year_path, month_dir)
                            if os.path.isdir(month_path):
                                for file in os.listdir(month_path):
                                    if file.startswith('Detalhado_') and file.endswith('.xlsx'):
                                        companies.add(company_dir)
                                        break
                            if company_dir in companies:
                                break
                    if company_dir in companies:
                        break
        
        return sorted(list(companies))
    
    def find_available_periods(self, company):
        """Encontra períodos disponíveis para uma empresa específica"""
        periods = set()
        
        abst_mot_dir = os.path.join(self.OUTPUT_BASE_DIR, 'Abst_Mot_Por_empresa', company)
        if not os.path.exists(abst_mot_dir):
            return []
        
        for year_dir in os.listdir(abst_mot_dir):
            year_path = os.path.join(abst_mot_dir, year_dir)
            if os.path.isdir(year_path):
                for month_dir in os.listdir(year_path):
                    month_path = os.path.join(year_path, month_dir)
                    if os.path.isdir(month_path):
                        for file in os.listdir(month_path):
                            if file.startswith('Detalhado_') and file.endswith('.xlsx'):
                                # Extrair mês e ano do nome do arquivo
                                parts = file.replace('.xlsx', '').split('_')
                                if len(parts) >= 3:
                                    month = parts[1]
                                    year = parts[2]
                                    periods.add(f"{month}_{year}")
        
        return sorted(list(periods))
    
    def determinar_turno(self, hora):
        """Determina o turno baseado na hora"""
        if isinstance(hora, str):
            try:
                hora_int = int(hora.split(':')[0])
            except:
                return None
        elif isinstance(hora, (datetime, pd.Timestamp)):
            hora_int = hora.hour
        else:
            return None
        
        for turno, definicao in self.turnos_definicao.items():
            if definicao['inicio'] <= hora_int <= definicao['fim']:
                return turno
        
        return None
    
    def calcular_tempo_por_turno(self, inicio, fim):
        """
        Calcula o tempo gasto em cada turno durante o período de trabalho
        Retorna um dicionário com o tempo em minutos para cada turno
        """
        if not isinstance(inicio, datetime) or not isinstance(fim, datetime):
            return {}
        
        # Ajustar para o próximo dia se necessário
        if fim < inicio:
            fim = fim + pd.Timedelta(days=1)
        
        tempo_por_turno = {turno: 0 for turno in self.turnos_definicao.keys()}
        
        # Calcular tempo total em minutos
        tempo_total = (fim - inicio).total_seconds() / 60
        
        if tempo_total <= 0:
            return tempo_por_turno
        
        # Obter a data base (sem hora) para cálculos
        data_base = inicio.date()
        
        # Processar cada turno
        for turno, definicao in self.turnos_definicao.items():
            # Definir início e fim do turno no dia
            turno_inicio_dia = datetime.combine(data_base, dtime(definicao['inicio'], 0, 0))
            turno_fim_dia = datetime.combine(data_base, dtime(definicao['fim'], 59, 59))
            
            # Calcular sobreposição entre o período de trabalho e o turno
            sobreposicao_inicio = max(inicio, turno_inicio_dia)
            sobreposicao_fim = min(fim, turno_fim_dia)
            
            # Se há sobreposição, calcular o tempo
            if sobreposicao_inicio < sobreposicao_fim:
                tempo_turno = (sobreposicao_fim - sobreposicao_inicio).total_seconds() / 60
                tempo_por_turno[turno] += tempo_turno
            
            # Verificar se o trabalho atravessa para o próximo dia
            if fim > datetime.combine(data_base, dtime(23, 59, 59)):
                # Calcular tempo no turno do próximo dia
                prox_dia = data_base + pd.Timedelta(days=1)
                turno_inicio_prox_dia = datetime.combine(prox_dia, dtime(definicao['inicio'], 0, 0))
                turno_fim_prox_dia = datetime.combine(prox_dia, dtime(definicao['fim'], 59, 59))
                
                # Calcular sobreposição no próximo dia
                sobreposicao_inicio_prox = max(fim, turno_inicio_prox_dia)
                sobreposicao_fim_prox = min(fim, turno_fim_prox_dia)
                
                # Se há sobreposição no próximo dia, calcular o tempo
                if sobreposicao_inicio_prox < sobreposicao_fim_prox:
                    tempo_turno_prox_dia = (sobreposicao_fim_prox - sobreposicao_inicio_prox).total_seconds() / 60
                    tempo_por_turno[turno] += tempo_turno_prox_dia
        
        return tempo_por_turno
    
    def process_company_period(self, company, month_year):
        """Processa um período específico de uma empresa"""
        try:
            month, year = month_year.split('_')
            
            # Caminho do arquivo Detalhado
            detalhado_path = os.path.join(
                self.OUTPUT_BASE_DIR, 
                'Abst_Mot_Por_empresa', 
                company, 
                year, 
                month.zfill(2),
                f'Detalhado_{company}_{month_year}{self.version_suffix}.xlsx'
            )
            
            if not os.path.exists(detalhado_path):
                logging.error(f"Arquivo Detalhado não encontrado: {detalhado_path}")
                return None
            
            logging.info(f"Processando Turnos Integração para {company} - {month_year}")
            
            # Carregar dados detalhados
            df_detalhado = pd.read_excel(detalhado_path)
            
            # Verificar colunas necessárias
            colunas_necessarias = ['motorista', 'matricula', 'placa', 'dia', 'pegada', 'largada', 'km_distributed', 'liters_distributed']
            colunas_faltantes = [col for col in colunas_necessarias if col not in df_detalhado.columns]
            
            if colunas_faltantes:
                logging.error(f"Colunas necessárias não encontradas: {colunas_faltantes}")
                return None
            
            # Converter colunas de data/hora
            df_detalhado['dia'] = pd.to_datetime(df_detalhado['dia'], format='%d/%m/%Y', errors='coerce')
            
            # Converter pegada e largada para datetime
            df_detalhado['pegada_dt'] = pd.to_datetime(df_detalhado['pegada'], format='%d/%m/%Y %H:%M', errors='coerce')
            df_detalhado['largada_dt'] = pd.to_datetime(df_detalhado['largada'], format='%d/%m/%Y %H:%M', errors='coerce')
            
            # Filtrar registros com horários válidos
            df_valido = df_detalhado[
                df_detalhado['pegada_dt'].notna() & 
                df_detalhado['largada_dt'].notna() &
                df_detalhado['km_distributed'].notna() &
                df_detalhado['liters_distributed'].notna()
            ].copy()
            
            if df_valido.empty:
                logging.warning(f"Nenhum registro válido encontrado para {company} - {month_year}")
                return None
            
            logging.info(f"Processando {len(df_valido)} registros válidos")
            
            # Lista para armazenar resultados
            resultados_turnos = []
            
            for idx, row in df_valido.iterrows():
                motorista = row['motorista']
                matricula = row['matricula']
                placa = row['placa']
                dia = row['dia']
                inicio = row['pegada_dt']
                fim = row['largada_dt']
                km_total = row['km_distributed']
                litros_total = row['liters_distributed']
                
                # Calcular tempo por turno
                tempo_por_turno = self.calcular_tempo_por_turno(inicio, fim)
                
                # Calcular tempo total
                tempo_total = sum(tempo_por_turno.values())
                
                if tempo_total > 0:
                    # Obter a data base para cálculos
                    data_base = inicio.date()
                    
                    # Distribuir km e litros proporcionalmente
                    for turno, tempo_turno in tempo_por_turno.items():
                        if tempo_turno > 0:
                            # Calcular proporção
                            proporcao = tempo_turno / tempo_total
                            
                            # Distribuir valores
                            km_turno = km_total * proporcao
                            litros_turno = litros_total * proporcao
                            
                            # Calcular horários específicos do turno
                            definicao_turno = self.turnos_definicao[turno]
                            turno_inicio_dia = datetime.combine(data_base, dtime(definicao_turno['inicio'], 0, 0))
                            turno_fim_dia = datetime.combine(data_base, dtime(definicao_turno['fim'], 59, 59))
                            
                            # Calcular sobreposição entre o período de trabalho e o turno
                            sobreposicao_inicio = max(inicio, turno_inicio_dia)
                            sobreposicao_fim = min(fim, turno_fim_dia)
                            
                            # Verificar se o trabalho atravessa para o próximo dia
                            if fim > datetime.combine(data_base, dtime(23, 59, 59)):
                                prox_dia = data_base + pd.Timedelta(days=1)
                                turno_inicio_prox_dia = datetime.combine(prox_dia, dtime(definicao_turno['inicio'], 0, 0))
                                turno_fim_prox_dia = datetime.combine(prox_dia, dtime(definicao_turno['fim'], 59, 59))
                                
                                sobreposicao_inicio_prox = max(fim, turno_inicio_prox_dia)
                                sobreposicao_fim_prox = min(fim, turno_fim_prox_dia)
                                
                                # Se há sobreposição no próximo dia, usar esses horários
                                if sobreposicao_inicio_prox < sobreposicao_fim_prox:
                                    hora_inicio_turno = sobreposicao_inicio_prox.strftime('%H:%M')
                                    hora_fim_turno = sobreposicao_fim_prox.strftime('%H:%M')
                                else:
                                    hora_inicio_turno = sobreposicao_inicio.strftime('%H:%M')
                                    hora_fim_turno = sobreposicao_fim.strftime('%H:%M')
                            else:
                                hora_inicio_turno = sobreposicao_inicio.strftime('%H:%M')
                                hora_fim_turno = sobreposicao_fim.strftime('%H:%M')
                            
                            # Adicionar ao resultado
                            resultados_turnos.append({
                                'motorista': motorista,
                                'matricula': matricula,
                                'placa': placa,
                                'dia': dia,
                                'turno': turno,
                                'hora_inicio_trabalho': hora_inicio_turno,
                                'hora_fim_trabalho': hora_fim_turno,
                                'tempo_turno_minutos': tempo_turno,
                                'tempo_total_minutos': tempo_total,
                                'proporcao_tempo': proporcao,
                                'km_distributed': km_turno,
                                'liters_distributed': litros_turno,
                                'km_l_turno': km_turno / litros_turno if litros_turno > 0 else 0
                            })
            
            if not resultados_turnos:
                logging.warning(f"Nenhum resultado de turno gerado para {company} - {month_year}")
                return None
            
            # Criar DataFrame final
            df_resultado = pd.DataFrame(resultados_turnos)
            
            # Arredondar valores
            df_resultado['km_distributed'] = df_resultado['km_distributed'].round(2)
            df_resultado['liters_distributed'] = df_resultado['liters_distributed'].round(2)
            df_resultado['km_l_turno'] = df_resultado['km_l_turno'].round(2)
            df_resultado['proporcao_tempo'] = df_resultado['proporcao_tempo'].round(4)
            
            logging.info(f"Processamento concluído: {len(df_resultado)} registros de turnos gerados")
            
            return df_resultado
            
        except Exception as e:
            logging.error(f"Erro ao processar Turnos Integração para {company} {month_year}: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return None
    
    def create_report(self, df_resultado, company, month_year):
        """Cria o relatório de turnos integração"""
        try:
            if df_resultado is None or df_resultado.empty:
                logging.warning("Nenhum dado para criar relatório de turnos integração")
                return False
            
            month, year = month_year.split('_')
            
            # Criar diretório de saída
            output_folder = os.path.join(self.OUTPUT_BASE_DIR, 'Turnos Integração', company, year, month.zfill(2))
            os.makedirs(output_folder, exist_ok=True)
            
            # Nome do arquivo
            output_file = os.path.join(output_folder, f'Turnos_Integração_{company}_{month_year}{self.version_suffix}.xlsx')
            
            logging.info(f"Criando relatório de turnos integração: {output_file}")
            
            # Verificar se o arquivo está em uso
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    if os.path.exists(output_file):
                        with open(output_file, 'r+b') as f:
                            pass
                    break
                except PermissionError:
                    if attempt < max_attempts - 1:
                        logging.warning(f"Tentativa {attempt + 1}/{max_attempts}: Arquivo {output_file} está em uso. Aguardando...")
                        tm.sleep(2)
                        base_name = output_file.replace('.xlsx', '')
                        output_file = f"{base_name}_{attempt + 1}.xlsx"
                    else:
                        logging.error(f"Arquivo {output_file} está em uso após {max_attempts} tentativas.")
                        raise Exception(f"Arquivo em uso: {output_file}. Feche o arquivo no Excel e tente novamente.")
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Aba principal com todos os dados
                df_resultado.to_excel(writer, sheet_name='Todos_Turnos', index=False)
                
                # Aba consolidada por motorista e turno
                df_consolidado = df_resultado.groupby(['motorista', 'matricula', 'turno']).agg({
                    'km_distributed': 'sum',
                    'liters_distributed': 'sum',
                    'tempo_turno_minutos': 'sum',
                    'dia': 'nunique'
                }).reset_index()
                
                df_consolidado['km_l_turno'] = df_consolidado['km_distributed'] / df_consolidado['liters_distributed']
                df_consolidado['km_l_turno'] = df_consolidado['km_l_turno'].round(2)
                df_consolidado['km_distributed'] = df_consolidado['km_distributed'].round(2)
                df_consolidado['liters_distributed'] = df_consolidado['liters_distributed'].round(2)
                df_consolidado['tempo_turno_minutos'] = df_consolidado['tempo_turno_minutos'].round(0)
                
                df_consolidado.to_excel(writer, sheet_name='Consolidado_Motorista_Turno', index=False)
                
                # Aba consolidada por turno
                df_turno_consolidado = df_resultado.groupby('turno').agg({
                    'km_distributed': 'sum',
                    'liters_distributed': 'sum',
                    'tempo_turno_minutos': 'sum',
                    'motorista': 'nunique',
                    'dia': 'nunique'
                }).reset_index()
                
                df_turno_consolidado['km_l_turno'] = df_turno_consolidado['km_distributed'] / df_turno_consolidado['liters_distributed']
                df_turno_consolidado['km_l_turno'] = df_turno_consolidado['km_l_turno'].round(2)
                df_turno_consolidado['km_distributed'] = df_turno_consolidado['km_distributed'].round(2)
                df_turno_consolidado['liters_distributed'] = df_turno_consolidado['liters_distributed'].round(2)
                df_turno_consolidado['tempo_turno_minutos'] = df_turno_consolidado['tempo_turno_minutos'].round(0)
                
                df_turno_consolidado.to_excel(writer, sheet_name='Consolidado_Turno', index=False)
                
                # Aplicar formatação condicional
                self.aplicar_formatacao_turnos(writer)
            
            logging.info(f"Relatório de turnos integração gerado com sucesso: {output_file}")
            return output_file
            
        except Exception as e:
            logging.error(f"Erro ao criar relatório de turnos integração: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            return False
    
    def aplicar_formatacao_turnos(self, writer):
        """Aplica formatação condicional ao relatório de turnos"""
        try:
            workbook = writer.book
            
            # Estilos
            vermelho_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            branco_bold = Font(color='FFFFFF', bold=True)
            amarelo_claro = PatternFill(start_color='FFFFE0', end_color='FFFFE0', fill_type='solid')
            preto_negrito = Font(color='000000', bold=True)
            verde_claro = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            verde_escuro_negrito = Font(color='006400', bold=True)
            azul_claro = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
            azul_escuro_negrito = Font(color='000080', bold=True)
            
            # Aplicar formatação em todas as abas
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Encontrar índices das colunas
                header = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
                
                idx_km = header.index('km_distributed') if 'km_distributed' in header else None
                idx_litros = header.index('liters_distributed') if 'liters_distributed' in header else None
                idx_km_l = header.index('km_l_turno') if 'km_l_turno' in header else None
                idx_tempo = header.index('tempo_turno_minutos') if 'tempo_turno_minutos' in header else None
                
                # Formatar cabeçalho
                for col_idx, cell in enumerate(next(worksheet.iter_rows(min_row=1, max_row=1))):
                    cell.font = Font(bold=True)
                
                # Formatar linhas de dados
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    # Formatação para km < 50
                    if idx_km is not None and row[idx_km].value is not None:
                        try:
                            valor_km = float(str(row[idx_km].value).replace(',', '.'))
                            if valor_km < 50:
                                cell_km = row[idx_km]
                                cell_km.fill = vermelho_fill
                                cell_km.font = branco_bold
                        except (ValueError, TypeError):
                            continue
                    
                    # Formatação para km/l < 3.5
                    if idx_km_l is not None and row[idx_km_l].value is not None:
                        try:
                            valor_km_l = float(str(row[idx_km_l].value).replace(',', '.'))
                            if valor_km_l < 3.5:
                                cell_km_l = row[idx_km_l]
                                cell_km_l.fill = amarelo_claro
                                cell_km_l.font = preto_negrito
                        except (ValueError, TypeError):
                            continue
                    
                    # Formatação para tempo de turno
                    if idx_tempo is not None and row[idx_tempo].value is not None:
                        try:
                            tempo_val = float(str(row[idx_tempo].value).replace(',', '.'))
                            cell_tempo = row[idx_tempo]
                            
                            if tempo_val < 30:
                                cell_tempo.fill = vermelho_fill
                                cell_tempo.font = branco_bold
                            elif 30 <= tempo_val <= 120:
                                cell_tempo.fill = amarelo_claro
                                cell_tempo.font = preto_negrito
                            elif 121 <= tempo_val <= 240:
                                cell_tempo.fill = azul_claro
                                cell_tempo.font = azul_escuro_negrito
                            elif tempo_val > 240:
                                cell_tempo.fill = verde_claro
                                cell_tempo.font = verde_escuro_negrito
                        except (ValueError, TypeError):
                            continue
                
                logging.info(f"Formatação aplicada na aba: {sheet_name}")
                
        except Exception as e:
            logging.error(f"Erro ao aplicar formatação: {str(e)}")

# --- GUI Unificada --- 

class ResumoMotoristaClienteProcessor:
    def __init__(self, base_dir, output_base_dir, version_suffix=""):
        self.BASE_DIR = base_dir
        self.SUPPLY_FOLDER = os.path.join(base_dir, 'Integração_Abast')
        self.RESUMO_FOLDER = os.path.join(base_dir, 'Resumo_Motorista_Cliente')
        self.OUTPUT_BASE_DIR = output_base_dir
        self.version_suffix = version_suffix
        
    def find_available_companies(self):
        """Encontra empresas disponíveis baseado nos arquivos de resumo"""
        logging.info("Buscando empresas disponíveis para Resumo_Motorista_Cliente...")
        
        if not os.path.exists(self.RESUMO_FOLDER):
            logging.error(f"Pasta de resumo não encontrada: {self.RESUMO_FOLDER}")
            return []
        
        resumo_files = [f for f in os.listdir(self.RESUMO_FOLDER) if f.endswith('.xlsx')]
        companies = set()
        
        for file in resumo_files:
            # Extrai o nome da empresa do arquivo (ex: RMC_Amparo_Agosto_2025.xlsx -> Amparo)
            parts = file.split('_')
            if len(parts) >= 3 and parts[0] == 'RMC':  # Formato: RMC_Empresa_Periodo.xlsx
                company = parts[1]
                companies.add(company)
            elif len(parts) >= 2:  # Formato alternativo: Empresa_Periodo.xlsx
                company = parts[0]
                companies.add(company)
        
        return sorted(list(companies))
    
    def find_available_periods(self, company):
        """Encontra períodos disponíveis para uma empresa específica"""
        if not os.path.exists(self.RESUMO_FOLDER):
            return []
        
        # Procura por arquivos com formato RMC_Empresa_Periodo.xlsx
        resumo_files = [f for f in os.listdir(self.RESUMO_FOLDER) if f.startswith(f'RMC_{company}_') and f.endswith('.xlsx')]
        periods = []
        
        for file in resumo_files:
            # Extrai o período do arquivo (ex: RMC_Amparo_Agosto_2025.xlsx -> Agosto_2025)
            parts = file.replace('.xlsx', '').split('_')
            if len(parts) >= 4 and parts[0] == 'RMC':  # Formato: RMC_Empresa_Periodo.xlsx
                period = '_'.join(parts[2:])
                periods.append(period)
            elif len(parts) >= 3:  # Formato alternativo: Empresa_Periodo.xlsx
                period = '_'.join(parts[1:])
                periods.append(period)
        
        return sorted(periods)
    
    def extract_plate_number(self, plate):
        """Extrai o número da placa, removendo letras MAR, A e RJ"""
        if pd.isna(plate):
            return None
        
        placa_str = str(plate).strip().upper()
        
        # Extrai apenas números da placa
        numeros = re.findall(r'\d+', placa_str)
        if numeros:
            # Se há múltiplos números, tenta combinar (ex: 01.124 -> 1124)
            if len(numeros) > 1:
                # Combina os números (ex: ['01', '124'] -> 1124)
                numero_combinado = int(''.join(numeros))
                if numero_combinado >= 1000:
                    return numero_combinado
            
            # Pega o maior número encontrado
            numero = int(max(numeros, key=len))
            # Retorna apenas se >= 1000
            if numero >= 1000:
                return numero
        
        return None
    
    def filter_plates_1000_plus(self, df, plate_column):
        """Filtra placas com número >= 1000"""
        if plate_column not in df.columns:
            logging.error(f"Coluna '{plate_column}' não encontrada no DataFrame")
            return df
        
        # Extrai números das placas
        df['plate_number'] = df[plate_column].apply(self.extract_plate_number)
        
        # Log para debug
        valid_numbers = df['plate_number'].dropna()
        logging.info(f"Placas válidas extraídas: {len(valid_numbers)} de {len(df)}")
        if len(valid_numbers) > 0:
            logging.info(f"Exemplos de números extraídos: {valid_numbers.head().tolist()}")
        
        # Filtra placas com número >= 1000
        filtered_df = df[df['plate_number'] >= 1000].copy()
        
        # Remove a coluna auxiliar
        filtered_df = filtered_df.drop('plate_number', axis=1)
        
        logging.info(f"Filtradas {len(filtered_df)} placas com número >= 1000 de {len(df)} total")
        return filtered_df
    
    def calculate_abastecimento_metrics(self, company, month_year):
        """Calcula métricas do arquivo de abastecimento para veículos >= 1000"""
        abastecimento_file = f"Abastecimento_{company}_{month_year}.xlsx"
        abastecimento_path = os.path.join(self.SUPPLY_FOLDER, abastecimento_file)
        
        if not os.path.exists(abastecimento_path):
            logging.error(f"Arquivo de abastecimento não encontrado: {abastecimento_path}")
            return None
        
        try:
            # Lê o arquivo de abastecimento
            df_abast = pd.read_excel(abastecimento_path, engine='openpyxl')
            logging.info(f"Arquivo de abastecimento carregado: {len(df_abast)} registros")
            
            # Identifica colunas relevantes
            plate_col = None
            km_col = None
            litros_col = None
            
            logging.info(f"Colunas disponíveis no arquivo de abastecimento: {list(df_abast.columns)}")
            
            for col in df_abast.columns:
                col_lower = str(col).lower()
                if 'placa' in col_lower:
                    plate_col = col
                elif 'km' in col_lower and 'litro' not in col_lower:
                    km_col = col
                elif 'litro' in col_lower or 'l' in col_lower:
                    litros_col = col
            
            logging.info(f"Colunas identificadas - Placa: {plate_col}, KM: {km_col}, Litros: {litros_col}")
            
            if not all([plate_col, km_col, litros_col]):
                logging.error(f"Colunas necessárias não encontradas. Placa: {plate_col}, KM: {km_col}, Litros: {litros_col}")
                return None
            
            # Filtra placas >= 1000
            df_filtered = self.filter_plates_1000_plus(df_abast, plate_col)
            
            if len(df_filtered) == 0:
                logging.warning("Nenhuma placa >= 1000 encontrada no arquivo de abastecimento")
                return None
            
            # Calcula totais
            total_km = df_filtered[km_col].sum()
            total_litros = df_filtered[litros_col].sum()
            
            if total_litros == 0:
                logging.warning("Total de litros é zero, não é possível calcular Km/l")
                return None
            
            km_l_abast = total_km / total_litros
            
            metrics = {
                'total_km': total_km,
                'total_litros': total_litros,
                'km_l_abast': km_l_abast,
                'veiculos_count': len(df_filtered),
                'plate_column': plate_col,
                'km_column': km_col,
                'litros_column': litros_col
            }
            
            logging.info(f"Métricas calculadas - KM: {total_km:.2f}, Litros: {total_litros:.2f}, Km/l: {km_l_abast:.2f}")
            return metrics
            
        except Exception as e:
            logging.error(f"Erro ao processar arquivo de abastecimento: {str(e)}")
            return None
    
    def process_company_period(self, company, month_year):
        """Processa um período específico para uma empresa - Nova implementação RMC"""
        logging.info(f"Processando Resumo_Motorista_Cliente para {company} - {month_year}")
        
        # 1. Carrega o arquivo de resumo (formato: RMC_Empresa_Periodo.xlsx)
        resumo_file = f"RMC_{company}_{month_year}.xlsx"
        resumo_path = os.path.join(self.RESUMO_FOLDER, resumo_file)
        
        if not os.path.exists(resumo_path):
            logging.error(f"Arquivo de resumo não encontrado: {resumo_path}")
            return None
        
        try:
            df_resumo = pd.read_excel(resumo_path, engine='openpyxl')
            logging.info(f"Arquivo de resumo carregado: {len(df_resumo)} registros")
            logging.info(f"Estrutura original do arquivo: {list(df_resumo.columns)}")
            
            # 2. Processar abastecimento
            abast_file = f"Abastecimento_{company}_{month_year}.xlsx"
            abast_path = os.path.join(self.SUPPLY_FOLDER, abast_file)
            
            if not os.path.exists(abast_path):
                logging.error(f"Arquivo de abastecimento não encontrado: {abast_path}")
                return None
            
            df_abast = pd.read_excel(abast_path, engine='openpyxl')
            logging.info(f"Arquivo de abastecimento carregado: {len(df_abast)} registros")
            
            # Limpar placas e filtrar abastecimento
            df_abast['placa_limpa'] = df_abast['placa'].apply(self.extract_plate_number)
            df_abast_filtrado = df_abast[df_abast['placa_limpa'].notna()].copy()
            logging.info(f"Placas válidas no abastecimento: {len(df_abast_filtrado)}")
            
            # Consolidar por veículo
            df_consolidado = df_abast_filtrado.groupby('placa_limpa').agg({
                'km': 'sum',
                'litros': 'sum'
            }).reset_index()
            
            df_consolidado['km_l_integracao'] = df_consolidado['km'] / df_consolidado['litros']
            df_consolidado = df_consolidado.rename(columns={
                'placa_limpa': 'placa_numero',
                'km': 'km_total',
                'litros': 'litros_total'
            })
            
            logging.info(f"Veículos consolidados: {len(df_consolidado)}")
            
            # 3. Limpar placas do RMC
            df_resumo['placa_limpa'] = df_resumo['placa'].apply(self.extract_plate_number)
            
            # 4. Fazer merge
            df_merged = df_resumo.merge(
                df_consolidado,
                left_on='placa_limpa',
                right_on='placa_numero',
                how='left'
            )
            
            registros_com_dados = df_merged['km_total'].notna().sum()
            logging.info(f"Registros com dados de abastecimento: {registros_com_dados}")
            
            # 5. Calcular distribuição proporcional baseada nos horários de início e fim
            # Para cada veículo, distribuir os totais de abastecimento proporcionalmente aos horários de trabalho
            
            # Extrair apenas a parte do horário (últimos 5 caracteres)
            df_merged['inicio_hora'] = df_merged['inicio'].astype(str).str[-5:]
            df_merged['fim_hora'] = df_merged['fim'].astype(str).str[-5:]
            
            # Converter horários para datetime para calcular duração
            df_merged['inicio_dt'] = pd.to_datetime(df_merged['inicio_hora'], format='%H:%M', errors='coerce')
            df_merged['fim_dt'] = pd.to_datetime(df_merged['fim_hora'], format='%H:%M', errors='coerce')
            
            # Calcular duração em horas para cada registro
            df_merged['duracao_horas'] = (df_merged['fim_dt'] - df_merged['inicio_dt']).dt.total_seconds() / 3600
            
            # Tratar casos onde fim < início (turno noturno)
            mask_turno_noturno = df_merged['fim_dt'] < df_merged['inicio_dt']
            df_merged.loc[mask_turno_noturno, 'duracao_horas'] = (df_merged.loc[mask_turno_noturno, 'fim_dt'] + pd.Timedelta(days=1) - df_merged.loc[mask_turno_noturno, 'inicio_dt']).dt.total_seconds() / 3600
            
            # Calcular total de horas trabalhadas por veículo
            df_merged['total_horas_veiculo'] = df_merged.groupby('placa_limpa')['duracao_horas'].transform('sum')
            
            # Calcular proporção de horas para cada registro
            df_merged['proporcao_horas'] = df_merged['duracao_horas'] / df_merged['total_horas_veiculo']
            
            # Distribuir os totais de abastecimento proporcionalmente às horas trabalhadas
            df_merged['km_distribuido'] = df_merged['km_total'] * df_merged['proporcao_horas']
            df_merged['litros_distribuido'] = df_merged['litros_total'] * df_merged['proporcao_horas']
            df_merged['km_l_distribuido'] = df_merged['km_distribuido'] / df_merged['litros_distribuido']
            
            # 6. Reorganizar colunas
            colunas_originais = ['matricula', 'nome', 'fase', 'placa', 'linha', 'inicio', 'fim', 
                                'km', 'lts', 'km/l', 'giro', 'freio', 'pedal', 'h/e', 'dia', 'app', 'dias']
            colunas_adicionais = ['placa_numero', 'km_total', 'litros_total', 'km_l_integracao',
                                 'duracao_horas', 'total_horas_veiculo', 'proporcao_horas',
                                 'km_distribuido', 'litros_distribuido', 'km_l_distribuido']
            
            colunas_finais = [col for col in colunas_originais + colunas_adicionais if col in df_merged.columns]
            df_final = df_merged[colunas_finais]
            
            # 7. Relatório final
            logging.info(f"Processamento concluído: {len(df_final)} registros")
            logging.info(f"Registros com dados distribuídos: {registros_com_dados}")
            
            if registros_com_dados > 0:
                km_l_original = df_final['km/l'].mean()
                km_l_distribuido = df_final['km_l_distribuido'].mean()
                km_l_integracao = df_final['km_l_integracao'].mean()
                
                logging.info(f"Km/l original RMC: {km_l_original:.2f}")
                logging.info(f"Km/l distribuído: {km_l_distribuido:.2f}")
                logging.info(f"Km/l integração: {km_l_integracao:.2f}")
            
            return df_final
            
        except Exception as e:
            logging.error(f"Erro ao processar arquivo de resumo: {str(e)}")
            return None
    
    def create_report(self, df_final, company, month_year):
        """Cria o relatório final RMC_Km/l_Distribuida"""
        if df_final is None or len(df_final) == 0:
            logging.error("DataFrame vazio, não é possível criar relatório")
            return False
        
        try:
            # Cria diretório de saída no local correto
            output_folder = os.path.join(self.OUTPUT_BASE_DIR, "RMC_Destribuida", company, "2025", month_year.split('_')[0])
            os.makedirs(output_folder, exist_ok=True)
            
            # Nome do arquivo de saída
            filename = f"RMC_Km_l_Distribuida_{company}_{month_year}{self.version_suffix}.xlsx"
            filepath = os.path.join(output_folder, filename)
            
            # Cria o arquivo Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df_final.to_excel(writer, sheet_name='Sheet1', index=False)
            
            logging.info(f"Relatório RMC_Km/l_Distribuida criado com sucesso: {filepath}")
            logging.info(f"Estrutura: {len(df_final.columns)} colunas, {len(df_final)} registros")
            logging.info(f"Colunas: {list(df_final.columns)}")
            return True
            
        except Exception as e:
            logging.error(f"Erro ao criar relatório: {str(e)}")
            return False


class UnifiedProcessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Processador Unificado de Relatórios")
        self.root.geometry("820x900")
        
        # Configuração responsiva da raiz
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        
        # Configuração de Tema (Auto Dark/Light)
        try:
            # Detecta se o sistema está em modo escuro
            if darkdetect.isDark():
                sv_ttk.set_theme("dark")
            else:
                sv_ttk.set_theme("light")
            
            # Hack opcional: Força a atualização da barra de título no Windows 11
            # (sv_ttk geralmente faz isso, mas garante a consistência)
        except Exception as e:
            logging.warning(f"Não foi possível aplicar o tema moderno: {e}")
        
        # Caminhos padrão - modifique aqui para definir caminhos pré-preenchidos
        self.DEFAULT_INPUT_DIR = r"D:\Scripts\Entrada"
        self.DEFAULT_OUTPUT_DIR = r"D:\Scripts\Saida"
        
        # Criar pastas se não existirem
        if self.DEFAULT_INPUT_DIR and not os.path.exists(self.DEFAULT_INPUT_DIR):
            os.makedirs(self.DEFAULT_INPUT_DIR, exist_ok=True)
            logging.info(f"Pasta de entrada criada: {self.DEFAULT_INPUT_DIR}")
        
        if self.DEFAULT_OUTPUT_DIR and not os.path.exists(self.DEFAULT_OUTPUT_DIR):
            os.makedirs(self.DEFAULT_OUTPUT_DIR, exist_ok=True)
            logging.info(f"Pasta de saída criada: {self.DEFAULT_OUTPUT_DIR}")
        
        self.base_dir = self.DEFAULT_INPUT_DIR
        self.output_base_dir = self.DEFAULT_OUTPUT_DIR
        self.company_processor = None
        self.ranking_processor = None
        self.ranking_integracao_processor = None # Novo processador
        self.ranking_ouro_mediano_processor = None # Novo processador para consolidação Ouro Mediano
        self.ranking_km_proporcional_processor = None # Novo processador para Ranking_Km_Proporcional
        self.turnos_integracao_processor = None # Novo processador para Turnos Integração
        self.company_months_abst = {}  # Meses disponíveis para Abst_Mot_Por_empresa
        self.company_periods_ranking = {} # Períodos disponíveis para Ranking_Por_Empresa
        self.company_periods_ranking_integracao = {} # Períodos disponíveis para Ranking_Integração
        self.company_periods_ranking_ouro_mediano = {} # Períodos disponíveis para Ranking_Ouro_Mediano
        self.company_periods_ranking_km_proporcional = {} # Períodos disponíveis para Ranking_Km_Proporcional
        self.company_periods_turnos_integracao = {} # Períodos disponíveis para Turnos Integração
        self.company_periods_resumo_motorista_cliente = {} # Períodos disponíveis para Resumo_Motorista_Cliente
        
        # Variáveis para controle de progresso
        self.total_tasks = 0
        self.completed_tasks = 0
        self.current_task = ""
        
        self.create_widgets()
        # Ajuste automático do tamanho da janela após criar widgets
        # self.root.update_idletasks()
        # self.root.geometry("")
    
    def create_widgets(self):
        # Frame principal com scroll
        main_frame = ttk.Frame(self.root, padding=10)
        main_frame.grid(row=0, column=0, sticky="nsew")
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(0, weight=1)
        
        # Canvas para scroll
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        # Configurar coluna 0 do scrollable_frame com weight=1 para esticar horizontalmente
        scrollable_frame.columnconfigure(0, weight=1)
        # Configurar rows para expansão vertical
        scrollable_frame.rowconfigure(4, weight=1)  # selection_frame (empresas e períodos)
        scrollable_frame.rowconfigure(7, weight=1)  # log_frame
        
        # Armazenar o ID da janela do canvas
        self.canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        
        # Função para ajustar largura e scroll region
        def configure_scroll_region(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            canvas.itemconfig(self.canvas_window, width=event.width)
        
        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", configure_scroll_region)
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Diretório Base
        dir_frame = ttk.LabelFrame(scrollable_frame, text="Diretório Base dos Arquivos de Entrada", padding=10)
        dir_frame.grid(row=0, column=0, sticky="ew", pady=5)
        dir_frame.columnconfigure(1, weight=1)
        dir_frame.rowconfigure(0, weight=1)
        
        # Label (coluna 0)
        dir_label = ttk.Label(dir_frame, text="Caminho:")
        dir_label.grid(row=0, column=0, padx=(0, 5), sticky="w", pady=10)
        
        # Entry (coluna 1) - responsivo
        self.dir_entry = ttk.Entry(dir_frame)
        # Pré-preenchido com o caminho padrão definido
        self.dir_entry.insert(0, self.base_dir)
        self.dir_entry.grid(row=0, column=1, sticky="ew", padx=(0, 5), pady=10)
        
        # Botão (coluna 2)
        browse_btn = ttk.Button(dir_frame, text="Procurar", command=self.browse_directory)
        browse_btn.grid(row=0, column=2, sticky="e", padx=5, pady=10)

        # Diretório de Saída
        output_dir_frame = ttk.LabelFrame(scrollable_frame, text="Diretório Base dos Arquivos de Saída", padding=10)
        output_dir_frame.grid(row=1, column=0, sticky="ew", pady=5)
        output_dir_frame.columnconfigure(1, weight=1)
        output_dir_frame.rowconfigure(0, weight=1)
        
        # Label (coluna 0)
        output_dir_label = ttk.Label(output_dir_frame, text="Caminho:")
        output_dir_label.grid(row=0, column=0, padx=(0, 5), sticky="w", pady=10)
        
        # Entry (coluna 1) - responsivo
        self.output_dir_entry = ttk.Entry(output_dir_frame)
        self.output_dir_entry.insert(0, self.output_base_dir)
        self.output_dir_entry.grid(row=0, column=1, sticky="ew", padx=(0, 5), pady=10)
        
        # Botão (coluna 2)
        browse_output_btn = ttk.Button(output_dir_frame, text="Procurar", command=self.browse_output_directory)
        browse_output_btn.grid(row=0, column=2, sticky="e", padx=5, pady=10)

        # Campo de Versão
        version_frame = ttk.LabelFrame(scrollable_frame, text="Versão dos Arquivos (ex: _2.0, _1.0)", padding=10)
        version_frame.grid(row=2, column=0, sticky="ew", pady=5)
        self.version_entry = ttk.Entry(version_frame, width=20)
        self.version_entry.insert(0, "")  # Inicia em branco
        self.version_entry.grid(row=0, column=0, padx=(0, 10), sticky="w")
        self.version_entry.bind('<KeyRelease>', self.on_version_change)
        
        # Dropdown de versões predefinidas
        self.version_options = ["---", "_1.0", "_2.0", "_3.0", "_4.0", "_5.0"]
        self.version_combobox = ttk.Combobox(version_frame, values=self.version_options, state="readonly", width=10)
        self.version_combobox.set("---")  # Valor inicial
        self.version_combobox.grid(row=0, column=1, padx=(0, 10), sticky="w")
        self.version_combobox.bind('<<ComboboxSelected>>', self.on_version_dropdown_select)

        # Seleção do Tipo de Relatório
        report_type_frame = ttk.LabelFrame(scrollable_frame, text="Tipo de Relatório", padding=10)
        report_type_frame.grid(row=3, column=0, sticky="ew", pady=5)
        
        # Checkbuttons para seleção múltipla
        self.report_types = ["Abst_Mot_Por_empresa", "Ranking_Por_Empresa", "Ranking_Integração", "Ranking_Ouro_Mediano", "Ranking_Km_Proporcional", "Turnos_Integração", "Resumo_Motorista_Cliente"]
        self.report_type_vars = {rt: tk.BooleanVar(value=True) for rt in self.report_types}
        
        # Configuração do Grid (3 Colunas)
        cols_per_row = 3
        total_items = len(self.report_types)
        rows_needed = (total_items + cols_per_row - 1) // cols_per_row  # Calcula total de linhas necessárias
        
        for i, rt in enumerate(self.report_types):
            # Calcula posição (linha, coluna) baseada no índice
            r, c = divmod(i, cols_per_row)
            ttk.Checkbutton(
                report_type_frame, 
                text=rt, 
                variable=self.report_type_vars[rt], 
                command=self.on_report_type_change
            ).grid(row=r, column=c, padx=10, pady=2, sticky="w")
            
        # ### LÓGICA DE ESPAÇAMENTO E BOTÃO ###
        
        # Coluna de espaçamento (Spacer): Coluna logo após as colunas de dados (índice 3)
        # Ela ganha weight=1 para ocupar todo o espaço vazio horizontal
        spacer_col = cols_per_row
        report_type_frame.columnconfigure(spacer_col, weight=1) 
        
        # Botão Processar Tudo
        # Posicionado na coluna após o spacer.
        # Rowspan define que ele ocupará a altura de todas as linhas de checkboxes geradas.
        self.process_everything_btn = ttk.Button(report_type_frame, text="Processar Tudo", command=self.process_everything)
        self.process_everything_btn.grid(row=0, column=spacer_col + 1, rowspan=rows_needed, padx=(10, 0), pady=10, sticky="e")

        # Frame para seleção de empresas e períodos
        selection_frame = ttk.Frame(scrollable_frame)
        selection_frame.grid(row=4, column=0, sticky="nsew", pady=5)
        selection_frame.columnconfigure(0, weight=1)
        selection_frame.columnconfigure(1, weight=1)
        selection_frame.rowconfigure(0, weight=1)
        
        # Seleção da Empresa
        company_frame = ttk.LabelFrame(selection_frame, text="Empresa", padding=10)
        company_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        company_frame.columnconfigure(0, weight=1)
        company_frame.rowconfigure(0, weight=1)
        # Listbox com seleção múltipla
        self.company_listbox = tk.Listbox(company_frame, selectmode=tk.MULTIPLE, exportselection=False)
        self.company_listbox.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.company_listbox.bind('<<ListboxSelect>>', self.on_company_select)
        self.company_scroll = ttk.Scrollbar(company_frame, orient=tk.VERTICAL, command=self.company_listbox.yview)
        self.company_scroll.grid(row=0, column=1, sticky="ns")
        self.company_listbox.config(yscrollcommand=self.company_scroll.set)

        # Seleção do Período (Ano/Mês)
        period_frame = ttk.LabelFrame(selection_frame, text="Períodos Disponíveis", padding=10)
        period_frame.grid(row=0, column=1, sticky="nsew")
        period_frame.columnconfigure(0, weight=1)
        period_frame.rowconfigure(0, weight=1)
        
        # Frame interno para organizar anos e meses
        period_inner_frame = ttk.Frame(period_frame)
        period_inner_frame.grid(row=0, column=0, sticky="nsew")
        period_inner_frame.columnconfigure(0, weight=1)
        period_inner_frame.columnconfigure(1, weight=1)
        period_inner_frame.rowconfigure(0, weight=1)
        
        # Frame para anos
        year_frame = ttk.LabelFrame(period_inner_frame, text="Anos", padding=5)
        year_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        year_frame.columnconfigure(0, weight=1)
        year_frame.rowconfigure(0, weight=1)
        
        # Listbox para ano
        self.year_listbox = tk.Listbox(year_frame, selectmode=tk.MULTIPLE, exportselection=False, height=6)
        self.year_listbox.grid(row=0, column=0, sticky="nsew", padx=(0, 2))
        year_scroll = ttk.Scrollbar(year_frame, orient=tk.VERTICAL, command=self.year_listbox.yview)
        year_scroll.grid(row=0, column=1, sticky="ns")
        self.year_listbox.config(yscrollcommand=year_scroll.set)
        
        # Frame para meses
        month_frame = ttk.LabelFrame(period_inner_frame, text="Meses", padding=5)
        month_frame.grid(row=0, column=1, sticky="nsew")
        month_frame.columnconfigure(0, weight=1)
        month_frame.rowconfigure(0, weight=1)
        
        # Listbox para mês
        self.month_listbox = tk.Listbox(month_frame, selectmode=tk.MULTIPLE, exportselection=False, height=6)
        self.month_listbox.grid(row=0, column=0, sticky="nsew", padx=(0, 2))
        month_scroll = ttk.Scrollbar(month_frame, orient=tk.VERTICAL, command=self.month_listbox.yview)
        month_scroll.grid(row=0, column=1, sticky="ns")
        self.month_listbox.config(yscrollcommand=month_scroll.set)

        # Botões de Ação
        button_frame = ttk.Frame(scrollable_frame, padding=10)
        button_frame.grid(row=5, column=0, sticky="ew", pady=5)
        
        # Configuração de colunas para os botões (0, 1, 2 tem peso igual para distribuição uniforme)
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=1)
        button_frame.columnconfigure(2, weight=1)
        
        # Botão 1 (Linha 0, Coluna 0)
        self.process_selected_btn = ttk.Button(button_frame, text="Processar Selecionados", state=tk.DISABLED, command=self.process_selected)
        self.process_selected_btn.grid(row=0, column=0, padx=5, pady=5, sticky="ew")
        
        # Botão 2 (Linha 0, Coluna 1)
        self.process_all_btn = ttk.Button(button_frame, text="Processar Todos os Períodos", state=tk.DISABLED, command=self.process_all_periods_for_company)
        self.process_all_btn.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        
        # Botão 3 (Linha 0, Coluna 2)
        self.process_all_companies_btn = ttk.Button(button_frame, text="Processar Todas as Empresas", command=self.process_all_companies)
        self.process_all_companies_btn.grid(row=0, column=2, padx=5, pady=5, sticky="ew")
        
        # Botão 4 (Linha 1, Coluna 0)
        self.process_ouro_mediano_btn = ttk.Button(button_frame, text="Consolidar Ouro Mediano", command=self.process_ouro_mediano_consolidation)
        self.process_ouro_mediano_btn.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        
        # Botão 5 (Linha 1, Coluna 1)
        self.process_km_proporcional_btn = ttk.Button(button_frame, text="Processar Ranking_Km_Proporcional", command=self.process_km_proporcional)
        self.process_km_proporcional_btn.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Botão 6 (Linha 1, Coluna 2) - Relatório Motoristas Insuficientes
        self.rpp_insuficientes_btn = ttk.Button(button_frame, text="Gerar Relatório Insuficientes", command=self.open_rpp_insuficientes_modal)
        self.rpp_insuficientes_btn.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        # ### LÓGICA DE ESPAÇAMENTO E BOTÃO DIREITO ###
        
        # Coluna de espaçamento (Spacer): Coluna 3 ganha todo o espaço vazio extra
        button_frame.columnconfigure(3, weight=10) 
        
        # Botão de Atualizar
        # Posicionado na Coluna 4. Rowspan=2 para ocupar a altura das duas linhas de botões à esquerda.
        self.refresh_btn = ttk.Button(button_frame, text="Atualizar", command=self.update_company_list)
        self.refresh_btn.grid(row=0, column=4, rowspan=2, padx=(20, 0), pady=5, sticky="e")
        
        # Frame de Progresso
        progress_frame = ttk.LabelFrame(scrollable_frame, text="Progresso", padding=10)
        progress_frame.grid(row=6, column=0, sticky="ew", pady=5)
        progress_frame.columnconfigure(0, weight=1)
        
        # Status atual
        self.status_var = tk.StringVar()
        self.status_var.set("Encontradas 18 empresas para os tipos de relatório selecionados.")
        # Status centralizado
        status_label = ttk.Label(progress_frame, textvariable=self.status_var, anchor=tk.CENTER)
        status_label.grid(row=0, column=0, sticky="ew", pady=(0, 5))
        
        # Barra de Progresso Geral
        self.progress = ttk.Progressbar(progress_frame, orient=tk.HORIZONTAL, mode='determinate')
        self.progress.grid(row=1, column=0, sticky="ew", pady=(0, 5))
        
        # Label de porcentagem centralizada
        self.progress_label = ttk.Label(progress_frame, text="0% (0/0)", anchor=tk.CENTER)
        self.progress_label.grid(row=2, column=0, pady=(0, 5))
        
        # Log de Processamento
        log_frame = ttk.LabelFrame(scrollable_frame, text="Log de Processamento", padding=10)
        log_frame.grid(row=7, column=0, sticky="nsew", pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        # Criar widget Text com scrollbar para o log
        log_text_frame = ttk.Frame(log_frame)
        log_text_frame.grid(row=0, column=0, sticky="nsew")
        log_text_frame.columnconfigure(0, weight=1)
        log_text_frame.rowconfigure(0, weight=1)
        
        # Text widget para o log
        self.log_text = tk.Text(log_text_frame, height=15, state=tk.DISABLED, 
                               wrap=tk.WORD, font=("Consolas", 9))
        self.log_text.grid(row=0, column=0, sticky="nsew")
        
        # Scrollbar para o log
        log_scrollbar = ttk.Scrollbar(log_text_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        log_scrollbar.grid(row=0, column=1, sticky="ns")
        self.log_text.config(yscrollcommand=log_scrollbar.set)
        
        # Configurar tags para colorir o log
        self.log_text.tag_config("info", foreground="blue")
        self.log_text.tag_config("success", foreground="green")
        self.log_text.tag_config("warning", foreground="orange")
        self.log_text.tag_config("error", foreground="red")
        self.log_text.tag_config("header", foreground="purple", font=("Consolas", 9, "bold"))
        
        # Botão para limpar o log
        clear_log_btn = ttk.Button(log_frame, text="Limpar Log", command=self.clear_log)
        clear_log_btn.grid(row=1, column=0, sticky="e", pady=(5, 0))
        
        # Botão para gerar PDF do relatório
        generate_pdf_btn = ttk.Button(log_frame, text="Gerar PDF do Relatório", command=self.generate_pdf_report)
        generate_pdf_btn.grid(row=2, column=0, sticky="e", pady=(5, 0))
        
        # Configurar scroll - usar grid
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        
        # Bind mouse wheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

    def browse_directory(self):
        # Obtém o caminho atual do campo ou usa um caminho padrão
        current_path = self.dir_entry.get().strip()
        # Normaliza o caminho para garantir compatibilidade entre discos (C:, D:, etc.)
        if current_path:
            current_path = os.path.normpath(current_path)
        if not current_path or not os.path.exists(current_path):
            # Se não houver caminho válido, usa o diretório atual ou um padrão
            current_path = os.path.normpath(os.getcwd())
        
        dir_path = filedialog.askdirectory(
            title="Selecione o diretório base dos arquivos de entrada",
            initialdir=current_path
        )
        if dir_path:
            # Normaliza o caminho selecionado para garantir compatibilidade entre discos
            dir_path = os.path.normpath(dir_path)
            self.dir_entry.delete(0, tk.END)
            self.dir_entry.insert(0, dir_path)
            self.base_dir = dir_path
            self.initialize_processors()
            self.update_company_list()
    
    def browse_output_directory(self):
        # Obtém o caminho atual do campo ou usa um caminho padrão
        current_path = self.output_dir_entry.get().strip()
        # Normaliza o caminho para garantir compatibilidade entre discos (C:, D:, etc.)
        if current_path:
            current_path = os.path.normpath(current_path)
        if not current_path or not os.path.exists(current_path):
            # Se não houver caminho válido, usa o diretório atual ou um padrão
            current_path = os.path.normpath(os.getcwd())
        
        dir_path = filedialog.askdirectory(
            title="Selecione o diretório base dos arquivos de saída",
            initialdir=current_path
        )
        if dir_path:
            # Normaliza o caminho selecionado para garantir compatibilidade entre discos
            dir_path = os.path.normpath(dir_path)
            self.output_dir_entry.delete(0, tk.END)
            self.output_dir_entry.insert(0, dir_path)
            self.output_base_dir = dir_path

    def initialize_processors(self):
        try:
            version_suffix = self.get_version_suffix()
            self.company_processor = CompanyProcessor(self.base_dir, self.output_base_dir, version_suffix)
            self.ranking_processor = RankingProcessor(self.base_dir, self.output_base_dir, version_suffix)
            self.ranking_integracao_processor = RankingIntegracaoProcessor(self.base_dir, self.output_base_dir, version_suffix) # Inicializa o novo processador
            self.ranking_ouro_mediano_processor = RankingOuroMedianoProcessor(self.base_dir, self.output_base_dir, version_suffix) # Inicializa o novo processador para consolidação Ouro Mediano
            self.ranking_km_proporcional_processor = RankingKmProporcionalProcessor(self.base_dir, self.output_base_dir, version_suffix) # Inicializa o novo processador para Ranking_Km_Proporcional
            self.turnos_integracao_processor = TurnosIntegracaoProcessor(self.base_dir, self.output_base_dir, version_suffix) # Inicializa o novo processador para Turnos Integração
            self.resumo_motorista_cliente_processor = ResumoMotoristaClienteProcessor(self.base_dir, self.output_base_dir, version_suffix) # Inicializa o novo processador para Resumo_Motorista_Cliente
            self.status_var.set(f"Diretórios base definidos: Entrada='{self.base_dir}', Saída='{self.output_base_dir}'")
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível inicializar os processadores:\n{str(e)}")
            self.status_var.set("Erro ao inicializar os processadores")

    def get_selected_report_types(self):
        return [rt for rt, var in self.report_type_vars.items() if var.get()]
    
    def get_version_suffix(self):
        """Retorna o sufixo de versão para os arquivos"""
        version = self.version_entry.get().strip()
        if version and not version.startswith('_'):
            version = '_' + version
        return version

    def update_company_list(self):
        # Reinicializa os processadores com a versão atual
        if self.base_dir and self.output_base_dir:
            self.initialize_processors()
            
        self.company_listbox.delete(0, tk.END)
        self.year_listbox.delete(0, tk.END)
        self.month_listbox.delete(0, tk.END)
        self.company_months_abst = {}
        self.company_periods_ranking = {}
        self.company_periods_ranking_integracao = {}
        self.company_periods_ranking_ouro_mediano = {}
        self.company_periods_ranking_km_proporcional = {}
        self.company_periods_turnos_integracao = {}
        self.company_periods_resumo_motorista_cliente = {}
        self.process_selected_btn.config(state=tk.DISABLED)
        self.process_all_btn.config(state=tk.DISABLED)

        selected_report_types = self.get_selected_report_types()
        companies_set = set()
        years_set = set()
        months_set = set()

        for report_type in selected_report_types:
            if report_type == "Abst_Mot_Por_empresa":
                if self.company_processor:
                    companies = self.company_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        file_pairs = self.company_processor.get_company_files(company)
                        periods = sorted(list(set([pair['month_year'] for pair in file_pairs])))
                        self.company_months_abst[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Ranking_Por_Empresa":
                if self.ranking_processor:
                    companies = self.ranking_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        periods = self.ranking_processor.find_available_periods(company)
                        self.company_periods_ranking[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Ranking_Integração":
                if self.ranking_integracao_processor:
                    companies = self.ranking_integracao_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        periods = self.ranking_integracao_processor.find_available_periods(company)
                        self.company_periods_ranking_integracao[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Ranking_Ouro_Mediano":
                if self.ranking_ouro_mediano_processor:
                    companies = self.ranking_ouro_mediano_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        periods = self.ranking_ouro_mediano_processor.find_available_periods(company)
                        self.company_periods_ranking_ouro_mediano[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Ranking_Km_Proporcional":
                if self.ranking_km_proporcional_processor:
                    companies = self.ranking_km_proporcional_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        periods = self.ranking_km_proporcional_processor.find_available_periods(company)
                        self.company_periods_ranking_km_proporcional[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Turnos_Integração":
                if self.turnos_integracao_processor:
                    companies = self.turnos_integracao_processor.find_available_companies()
                    for company in companies:
                        companies_set.add(company)
                        periods = self.turnos_integracao_processor.find_available_periods(company)
                        self.company_periods_turnos_integracao[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
            elif report_type == "Resumo_Motorista_Cliente":
                if self.resumo_motorista_cliente_processor:
                    logging.info("Processador Resumo_Motorista_Cliente encontrado, buscando empresas...")
                    companies = self.resumo_motorista_cliente_processor.find_available_companies()
                    logging.info(f"Empresas encontradas para Resumo_Motorista_Cliente: {companies}")
                    for company in companies:
                        companies_set.add(company)
                        periods = self.resumo_motorista_cliente_processor.find_available_periods(company)
                        logging.info(f"Períodos encontrados para {company}: {periods}")
                        self.company_periods_resumo_motorista_cliente[company] = periods
                        for period in periods:
                            m, y = period.split('_')
                            years_set.add(y)
                            months_set.add(m)
                else:
                    logging.error("Processador Resumo_Motorista_Cliente não foi inicializado!")

        companies_sorted = sorted(list(companies_set))
        for company in companies_sorted:
            self.company_listbox.insert(tk.END, company)

        # Separar anos e meses corretamente
        anos_limpos = set()
        meses_limpos = set()
        
        for item in years_set:
            # Se contém ponto, é um ano com versão (ex: 2025.1.0)
            if '.' in item:
                ano_base = item.split('.')[0]
                anos_limpos.add(ano_base)
            else:
                # Verificar se é realmente um ano (4 dígitos)
                if item.isdigit() and len(item) == 4:
                    anos_limpos.add(item)
        
        for item in months_set:
            # Verificar se é realmente um mês (não contém números)
            if not any(char.isdigit() for char in item):
                meses_limpos.add(item)

        # Inserir anos ordenados
        for y in sorted(anos_limpos):
            self.year_listbox.insert(tk.END, y)
        
        # Organizar meses em ordem cronológica
        meses_ordenados = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", 
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        
        # Filtrar apenas os meses que existem nos dados
        meses_disponiveis = [m for m in meses_ordenados if m in meses_limpos]
        
        for m in meses_disponiveis:
            self.month_listbox.insert(tk.END, m)

        if companies_sorted and years_set and months_set:
            self.process_selected_btn.config(state=tk.NORMAL)
            self.process_all_btn.config(state=tk.NORMAL)
        if not companies_sorted:
            self.status_var.set("Nenhuma empresa encontrada para os tipos de relatório selecionados.")
        else:
            self.status_var.set(f"Encontradas {len(companies_sorted)} empresas para os tipos de relatório selecionados.")

    def on_report_type_change(self, *args):
        self.update_company_list()
    
    def on_version_change(self, event=None):
        """Reinicializa os processadores quando a versão muda"""
        if self.base_dir and self.output_base_dir:
            self.initialize_processors()
            self.update_company_list()

    def on_version_dropdown_select(self, event=None):
        """Atualiza o campo de versão quando uma opção é selecionada no dropdown"""
        selected = self.version_combobox.get()
        if selected == "---":
            # Limpa o campo se selecionou o valor vazio
            self.version_entry.delete(0, tk.END)
        else:
            # Preenche o campo com a versão selecionada
            self.version_entry.delete(0, tk.END)
            self.version_entry.insert(0, selected)
        # Dispara a atualização dos processadores
        self.on_version_change()

    def on_company_select(self, event=None):
        # Atualiza os períodos disponíveis conforme empresas selecionadas
        selected_indices = self.company_listbox.curselection()
        selected_companies = [self.company_listbox.get(i) for i in selected_indices]
        selected_report_types = self.get_selected_report_types()
        years_set = set()
        months_set = set()
        for report_type in selected_report_types:
            for company in selected_companies:
                if report_type == "Abst_Mot_Por_empresa":
                    periods = self.company_months_abst.get(company, [])
                elif report_type == "Ranking_Por_Empresa":
                    periods = self.company_periods_ranking.get(company, [])
                elif report_type == "Ranking_Integração":
                    periods = self.company_periods_ranking_integracao.get(company, [])
                elif report_type == "Ranking_Ouro_Mediano":
                    periods = self.company_periods_ranking_ouro_mediano.get(company, [])
                elif report_type == "Ranking_Km_Proporcional":
                    periods = self.company_periods_ranking_km_proporcional.get(company, [])
                elif report_type == "Turnos_Integração":
                    periods = self.company_periods_turnos_integracao.get(company, [])
                elif report_type == "Resumo_Motorista_Cliente":
                    periods = self.company_periods_resumo_motorista_cliente.get(company, [])
                else:
                    periods = []
                for period in periods:
                    m, y = period.split('_')
                    years_set.add(y)
                    months_set.add(m)
        self.year_listbox.delete(0, tk.END)
        self.month_listbox.delete(0, tk.END)
        
        # Separar anos e meses corretamente
        anos_limpos = set()
        meses_limpos = set()
        
        for item in years_set:
            # Se contém ponto, é um ano com versão (ex: 2025.1.0)
            if '.' in item:
                ano_base = item.split('.')[0]
                anos_limpos.add(ano_base)
            else:
                # Verificar se é realmente um ano (4 dígitos)
                if item.isdigit() and len(item) == 4:
                    anos_limpos.add(item)
        
        for item in months_set:
            # Verificar se é realmente um mês (não contém números)
            if not any(char.isdigit() for char in item):
                meses_limpos.add(item)

        # Inserir anos ordenados
        for y in sorted(anos_limpos):
            self.year_listbox.insert(tk.END, y)
        
        # Organizar meses em ordem cronológica
        meses_ordenados = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", 
                          "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        
        # Filtrar apenas os meses que existem nos dados
        meses_disponiveis = [m for m in meses_ordenados if m in meses_limpos]
        
        for m in meses_disponiveis:
            self.month_listbox.insert(tk.END, m)
        if selected_companies and years_set and months_set:
            self.process_selected_btn.config(state=tk.NORMAL)
            self.process_all_btn.config(state=tk.NORMAL)
        else:
            self.process_selected_btn.config(state=tk.DISABLED)
            self.process_all_btn.config(state=tk.DISABLED)

    def get_selected_years_months(self):
        selected_years = [self.year_listbox.get(i) for i in self.year_listbox.curselection()]
        selected_months = [self.month_listbox.get(i) for i in self.month_listbox.curselection()]
        return selected_years, selected_months

    def process_selected(self):
        selected_company_indices = self.company_listbox.curselection()
        selected_companies = [self.company_listbox.get(i) for i in selected_company_indices]
        selected_report_types = self.get_selected_report_types()
        selected_years, selected_months = self.get_selected_years_months()
        if not selected_companies:
            messagebox.showwarning("Aviso", "Por favor, selecione pelo menos uma empresa.")
            return
        if not selected_report_types:
            messagebox.showwarning("Aviso", "Por favor, selecione pelo menos um tipo de relatório.")
            return
        if not selected_years or not selected_months:
            messagebox.showwarning("Aviso", "Por favor, selecione pelo menos um ano e um mês.")
            return
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        # Limpar log e iniciar processamento
        self.clear_log()
        self.add_log_entry("=== INICIANDO PROCESSAMENTO SELECIONADO ===", "header")
        self.add_log_entry(f"Empresas: {', '.join(selected_companies)}", "info")
        self.add_log_entry(f"Tipos de relatório: {', '.join(selected_report_types)}", "info")
        self.add_log_entry(f"Anos: {', '.join(selected_years)}", "info")
        self.add_log_entry(f"Meses: {', '.join(selected_months)}", "info")
        
        periods = [f"{m}_{y}" for y in selected_years for m in selected_months]
        for report_type in selected_report_types:
            for company in selected_companies:
                self.add_log_entry(f"Iniciando processamento: {company} [{report_type}]", "start")
                self.run_processing(report_type, company, periods)

    def process_all_periods_for_company(self):
        selected_company_indices = self.company_listbox.curselection()
        selected_companies = [self.company_listbox.get(i) for i in selected_company_indices]
        selected_report_types = self.get_selected_report_types()
        selected_years, selected_months = self.get_selected_years_months()
        if not selected_companies:
            messagebox.showwarning("Aviso", "Por favor, selecione pelo menos uma empresa.")
            return
        if not selected_report_types:
            messagebox.showwarning("Aviso", "Por favor, selecione pelo menos um tipo de relatório.")
            return
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        # Calcular total de tarefas para progresso
        total_tasks = 0
        for report_type in selected_report_types:
            for company in selected_companies:
                if report_type == "Abst_Mot_Por_empresa":
                    periods_to_process = self.company_months_abst.get(company, [])
                elif report_type == "Ranking_Por_Empresa":
                    periods_to_process = self.company_periods_ranking.get(company, [])
                elif report_type == "Ranking_Integração":
                    periods_to_process = self.company_periods_ranking_integracao.get(company, [])
                elif report_type == "Ranking_Ouro_Mediano":
                    periods_to_process = self.company_periods_ranking_ouro_mediano.get(company, [])
                elif report_type == "Ranking_Km_Proporcional":
                    periods_to_process = self.company_periods_ranking_km_proporcional.get(company, [])
                elif report_type == "Turnos_Integração":
                    periods_to_process = self.company_periods_turnos_integracao.get(company, [])
                elif report_type == "Resumo_Motorista_Cliente":
                    periods_to_process = self.company_periods_resumo_motorista_cliente.get(company, [])
                else:
                    periods_to_process = []
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods_to_process = [p for p in periods_to_process if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                total_tasks += len(periods_to_process)
        
        # Limpar log e iniciar processamento
        self.clear_log()
        self.add_log_entry("=== INICIANDO PROCESSAMENTO DE TODOS OS PERÍODOS ===", "header")
        
        # Inicializar progresso
        self.update_progress("Iniciando processamento...", 0, total_tasks)
        
        completed_tasks = 0
        for report_type in selected_report_types:
            for company in selected_companies:
                if report_type == "Abst_Mot_Por_empresa":
                    periods_to_process = self.company_months_abst.get(company, [])
                elif report_type == "Ranking_Por_Empresa":
                    periods_to_process = self.company_periods_ranking.get(company, [])
                elif report_type == "Ranking_Integração":
                    periods_to_process = self.company_periods_ranking_integracao.get(company, [])
                elif report_type == "Ranking_Ouro_Mediano":
                    periods_to_process = self.company_periods_ranking_ouro_mediano.get(company, [])
                elif report_type == "Ranking_Km_Proporcional":
                    periods_to_process = self.company_periods_ranking_km_proporcional.get(company, [])
                elif report_type == "Turnos_Integração":
                    periods_to_process = self.company_periods_turnos_integracao.get(company, [])
                elif report_type == "Resumo_Motorista_Cliente":
                    periods_to_process = self.company_periods_resumo_motorista_cliente.get(company, [])
                else:
                    periods_to_process = []
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods_to_process = [p for p in periods_to_process if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                
                if periods_to_process:
                    self.add_log_entry(f"Iniciando processamento: {company} [{report_type}] - {len(periods_to_process)} período(s)", "start")
                    self.run_processing(report_type, company, periods_to_process)
                    completed_tasks += len(periods_to_process)
                    self.update_progress(f"Concluído: {company} [{report_type}]", completed_tasks, total_tasks)
                    self.add_log_entry(f"Concluído: {company} [{report_type}] - {len(periods_to_process)} período(s) processados", "success")

    def process_all_companies(self):
        selected_report_types = self.get_selected_report_types()
        selected_years, selected_months = self.get_selected_years_months()
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        # Calcular total de tarefas para progresso
        total_tasks = 0
        for report_type in selected_report_types:
            if report_type == "Abst_Mot_Por_empresa":
                companies = list(self.company_months_abst.keys())
                periods_dict = self.company_months_abst
            elif report_type == "Ranking_Por_Empresa":
                companies = list(self.company_periods_ranking.keys())
                periods_dict = self.company_periods_ranking
            elif report_type == "Ranking_Integração":
                companies = list(self.company_periods_ranking_integracao.keys())
                periods_dict = self.company_periods_ranking_integracao
            elif report_type == "Ranking_Ouro_Mediano":
                companies = list(self.company_periods_ranking_ouro_mediano.keys())
                periods_dict = self.company_periods_ranking_ouro_mediano
            elif report_type == "Ranking_Km_Proporcional":
                companies = list(self.company_periods_ranking_km_proporcional.keys())
                periods_dict = self.company_periods_ranking_km_proporcional
            elif report_type == "Turnos_Integração":
                companies = list(self.company_periods_turnos_integracao.keys())
                periods_dict = self.company_periods_turnos_integracao
            elif report_type == "Resumo_Motorista_Cliente":
                companies = list(self.company_periods_resumo_motorista_cliente.keys())
                periods_dict = self.company_periods_resumo_motorista_cliente
            else:
                continue
            for company in companies:
                periods = periods_dict[company]
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods = [p for p in periods if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                total_tasks += len(periods)
        
        # Limpar log e iniciar processamento
        self.clear_log()
        self.add_log_entry("=== INICIANDO PROCESSAMENTO DE TODAS AS EMPRESAS ===", "header")
        
        # Inicializar progresso
        self.update_progress("Iniciando processamento...", 0, total_tasks)
        
        completed_tasks = 0
        for report_type in selected_report_types:
            self.add_log_entry(f"Processando tipo de relatório: {report_type}", "header")
            if report_type == "Abst_Mot_Por_empresa":
                companies = list(self.company_months_abst.keys())
                periods_dict = self.company_months_abst
            elif report_type == "Ranking_Por_Empresa":
                companies = list(self.company_periods_ranking.keys())
                periods_dict = self.company_periods_ranking
            elif report_type == "Ranking_Integração":
                companies = list(self.company_periods_ranking_integracao.keys())
                periods_dict = self.company_periods_ranking_integracao
            elif report_type == "Ranking_Ouro_Mediano":
                companies = list(self.company_periods_ranking_ouro_mediano.keys())
                periods_dict = self.company_periods_ranking_ouro_mediano
            elif report_type == "Ranking_Km_Proporcional":
                companies = list(self.company_periods_ranking_km_proporcional.keys())
                periods_dict = self.company_periods_ranking_km_proporcional
            elif report_type == "Turnos_Integração":
                companies = list(self.company_periods_turnos_integracao.keys())
                periods_dict = self.company_periods_turnos_integracao
            elif report_type == "Resumo_Motorista_Cliente":
                companies = list(self.company_periods_resumo_motorista_cliente.keys())
                periods_dict = self.company_periods_resumo_motorista_cliente
            else:
                continue
            for company in companies:
                periods = periods_dict[company]
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods = [p for p in periods if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                
                if periods:
                    self.add_log_entry(f"Iniciando: {company} [{report_type}] - {len(periods)} período(s)", "start")
                    self.run_processing(report_type, company, periods)
                    completed_tasks += len(periods)
                    self.update_progress(f"Concluído: {company} [{report_type}]", completed_tasks, total_tasks)
                    self.add_log_entry(f"Concluído: {company} [{report_type}] - {len(periods)} período(s) processados", "success")

    def process_everything(self):
        selected_report_types = self.get_selected_report_types()
        selected_years, selected_months = self.get_selected_years_months()
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        # Calcular total de tarefas para progresso
        total_tasks = 0
        for report_type in selected_report_types:
            if report_type == "Abst_Mot_Por_empresa":
                companies = list(self.company_months_abst.keys())
                periods_dict = self.company_months_abst
            elif report_type == "Ranking_Por_Empresa":
                companies = list(self.company_periods_ranking.keys())
                periods_dict = self.company_periods_ranking
            elif report_type == "Ranking_Integração":
                companies = list(self.company_periods_ranking_integracao.keys())
                periods_dict = self.company_periods_ranking_integracao
            elif report_type == "Ranking_Ouro_Mediano":
                companies = list(self.company_periods_ranking_ouro_mediano.keys())
                periods_dict = self.company_periods_ranking_ouro_mediano
            elif report_type == "Ranking_Km_Proporcional":
                companies = list(self.company_periods_ranking_km_proporcional.keys())
                periods_dict = self.company_periods_ranking_km_proporcional
            elif report_type == "Turnos_Integração":
                companies = list(self.company_periods_turnos_integracao.keys())
                periods_dict = self.company_periods_turnos_integracao
            elif report_type == "Resumo_Motorista_Cliente":
                companies = list(self.company_periods_resumo_motorista_cliente.keys())
                periods_dict = self.company_periods_resumo_motorista_cliente
            else:
                continue
            for company in companies:
                periods = periods_dict[company]
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods = [p for p in periods if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                total_tasks += len(periods)
        
        # Inicializar progresso
        self.update_progress("Iniciando processamento completo...", 0, total_tasks)
        
        completed_tasks = 0
        for report_type in selected_report_types:
            if report_type == "Abst_Mot_Por_empresa":
                companies = list(self.company_months_abst.keys())
                periods_dict = self.company_months_abst
            elif report_type == "Ranking_Por_Empresa":
                companies = list(self.company_periods_ranking.keys())
                periods_dict = self.company_periods_ranking
            elif report_type == "Ranking_Integração":
                companies = list(self.company_periods_ranking_integracao.keys())
                periods_dict = self.company_periods_ranking_integracao
            elif report_type == "Ranking_Ouro_Mediano":
                companies = list(self.company_periods_ranking_ouro_mediano.keys())
                periods_dict = self.company_periods_ranking_ouro_mediano
            elif report_type == "Ranking_Km_Proporcional":
                companies = list(self.company_periods_ranking_km_proporcional.keys())
                periods_dict = self.company_periods_ranking_km_proporcional
            elif report_type == "Turnos_Integração":
                companies = list(self.company_periods_turnos_integracao.keys())
                periods_dict = self.company_periods_turnos_integracao
            elif report_type == "Resumo_Motorista_Cliente":
                companies = list(self.company_periods_resumo_motorista_cliente.keys())
                periods_dict = self.company_periods_resumo_motorista_cliente
            else:
                continue
            for company in companies:
                periods = periods_dict[company]
                # Se anos/meses selecionados, filtra
                if selected_years and selected_months:
                    periods = [p for p in periods if any(p == f"{m}_{y}" for y in selected_years for m in selected_months)]
                
                if periods:
                    self.run_processing(report_type, company, periods)
                    completed_tasks += len(periods)
                    self.update_progress(f"Concluído: {company} [{report_type}]", completed_tasks, total_tasks)
        
        # Mensagem final
        self.status_var.set(f"🎉 Processamento completo finalizado! {completed_tasks}/{total_tasks} tarefas concluídas")

    def clear_log(self):
        """Limpa o log de processamento"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)
    
    def add_log_entry(self, message, level="info", log_level=None):
        """
        Adiciona uma entrada no log com timestamp, nível de log e formatação
        
        Args:
            message: Mensagem a ser exibida no log
            level: Nível customizado da interface ("info", "success", "error", "warning", "header", "start", "processing")
            log_level: Nível do logging padrão (logging.INFO, logging.WARNING, logging.ERROR, logging.DEBUG)
                      Se None, será mapeado automaticamente baseado no parâmetro 'level'
        """
        import datetime
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        
        # Mapear nível customizado para nível de logging se não fornecido
        if log_level is None:
            level_mapping = {
                "success": logging.INFO,
                "error": logging.ERROR,
                "warning": logging.WARNING,
                "header": logging.INFO,
                "start": logging.INFO,
                "processing": logging.INFO,
                "info": logging.INFO
            }
            log_level = level_mapping.get(level, logging.INFO)
        
        # Obter nome do nível de log para exibição
        log_level_name = logging.getLevelName(log_level)
        
        self.log_text.config(state=tk.NORMAL)
        
        # Adicionar timestamp e nível de log
        self.log_text.insert(tk.END, f"[{timestamp}] ", "info")
        self.log_text.insert(tk.END, f"[{log_level_name}] ", "info")
        
        # Determinar ícone e cor baseado no nível
        if level == "success":
            icon = "✅ "
            tag = "success"
        elif level == "error":
            icon = "❌ "
            tag = "error"
        elif level == "warning":
            icon = "⚠️ "
            tag = "warning"
        elif level == "header":
            icon = "📊 "
            tag = "header"
        elif level == "start":
            icon = "🚀 "
            tag = "info"
        elif level == "processing":
            icon = "⚙️ "
            tag = "info"
        else:
            icon = "ℹ️ "
            tag = "info"
        
        # Adicionar ícone e mensagem
        self.log_text.insert(tk.END, icon, tag)
        self.log_text.insert(tk.END, message + "\n", tag)
        
        # Scroll para o final
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        
        # Atualizar interface
        self.root.update_idletasks()
    
    def update_progress(self, current_task, completed=None, total=None):
        """Atualiza o progresso da interface"""
        if current_task:
            self.current_task = current_task
            self.status_var.set(f"Processando: {current_task}")
            # Adicionar ao log também
            if completed is not None and total is not None:
                self.add_log_entry(f"{current_task} ({completed}/{total})", "processing")
            else:
                self.add_log_entry(current_task, "processing")
        
        if completed is not None and total is not None:
            self.completed_tasks = completed
            self.total_tasks = total
            # Calcular porcentagem (limitada a 100%)
            percentage = min(100, (completed / total * 100)) if total > 0 else 0
            self.progress["value"] = percentage
            self.progress_label.config(text=f"{percentage:.1f}% ({completed}/{total})")
            
            # Resetar estilo da barra ao iniciar (completed=0)
            if completed == 0:
                try:
                    self.progress.configure(style="TProgressbar")  # Estilo padrão
                except Exception:
                    pass
            # Mudar estilo da barra para verde quando atingir 100%
            elif percentage >= 100:
                try:
                    style = ttk.Style()
                    style.configure("green.Horizontal.TProgressbar", troughcolor='gray', background='#88E788')
                    self.progress.configure(style="green.Horizontal.TProgressbar")
                except Exception:
                    pass  # Ignora se não conseguir mudar o estilo
        
        # Otimizado: atualizar UI apenas quando necessário (a cada 10 itens ou no final)
        if completed is not None and total is not None:
            # Atualizar UI apenas a cada 10% de progresso ou no último item
            if completed % max(1, total // 10) == 0 or completed == total:
                self.root.update_idletasks()
        else:
            self.root.update_idletasks()
    
    def run_processing(self, report_type, company, periods_to_process):
        total_periods = len(periods_to_process)
        success_count = 0
        
        start_time = tm.time()
        
        for i, period in enumerate(periods_to_process):
            current_task = f"{company} - {period} ({i+1}/{total_periods}) [{report_type}]"
            # Atualizar progresso (update_idletasks agora é otimizado dentro de update_progress)
            self.update_progress(current_task, i, total_periods)
            
            try:
                if report_type == "Abst_Mot_Por_empresa":
                    file_pairs = self.company_processor.get_company_files(company)
                    pair = next((p for p in file_pairs if p['month_year'] == period), None)
                    if pair:
                        success = self.company_processor.process_company_files(pair['supply'], pair['drivers'], company, pair['month_year'])
                    else:
                        logging.error(f"Par de arquivos não encontrado para {company} em {period}.")
                        success = False
                elif report_type == "Ranking_Por_Empresa":
                    df_result = self.ranking_processor.process_company_period(company, period)
                    if df_result is not None:
                        success = self.ranking_processor.create_report(df_result, company, period)
                    else:
                        success = False
                elif report_type == "Ranking_Integração":
                    df_result = self.ranking_integracao_processor.process_company_period(company, period)
                    if df_result is not None:
                        success = self.ranking_integracao_processor.create_report(df_result, company, period)
                    else:
                        success = False
                elif report_type == "Ranking_Ouro_Mediano":
                    # Para Ranking_Ouro_Mediano, processar consolidação de todos os períodos
                    df_result = self.ranking_ouro_mediano_processor.process_consolidation(
                        selected_companies=[company], 
                        selected_periods=periods_to_process
                    )
                    if df_result is not None:
                        success = self.ranking_ouro_mediano_processor.create_consolidated_report(df_result, selected_periods=periods_to_process)
                    else:
                        success = False
                elif report_type == "Ranking_Km_Proporcional":
                    result = self.ranking_km_proporcional_processor.process_company_period(company, period)
                    if result is not None:
                        success = True
                    else:
                        success = False
                elif report_type == "Turnos_Integração":
                    df_result = self.turnos_integracao_processor.process_company_period(company, period)
                    if df_result is not None:
                        success = self.turnos_integracao_processor.create_report(df_result, company, period)
                    else:
                        success = False
                elif report_type == "Resumo_Motorista_Cliente":
                    logging.info(f"Iniciando processamento Resumo_Motorista_Cliente para {company} - {period}")
                    if self.resumo_motorista_cliente_processor:
                        df_result = self.resumo_motorista_cliente_processor.process_company_period(company, period)
                        if df_result is not None:
                            logging.info(f"Processamento bem-sucedido, criando relatório para {company} - {period}")
                            success = self.resumo_motorista_cliente_processor.create_report(df_result, company, period)
                        else:
                            logging.error(f"Processamento retornou None para {company} - {period}")
                            success = False
                    else:
                        logging.error("Processador Resumo_Motorista_Cliente não foi inicializado!")
                        success = False
                
                if success:
                    success_count += 1
                    self.add_log_entry(f"✅ {company} - {period} processado com sucesso", "success")
                else:
                    self.add_log_entry(f"❌ Falha ao processar {company} - {period}", "error")
                    
            except Exception as e:
                logging.error(f"Erro crítico ao processar {company} - {period}: {traceback.format_exc()}")
                self.add_log_entry(f"❌ Erro crítico em {company} - {period}: {str(e)}", "error")
            
            # Pequeno delay para visualização da animação da barra de progresso
            tm.sleep(0.05)
            
            # Atualiza progresso após cada período
            self.update_progress(f"{company} - {period} [{report_type}]", i+1, total_periods)
            self.root.update_idletasks()
        
        # Registrar tempo de processamento
        processing_time = tm.time() - start_time
        logging.info(f"Tempo de processamento para {company}: {processing_time:.2f} segundos")
        
        # Atualizar progresso final
        self.update_progress(f"Concluído: {company} [{report_type}]", total_periods, total_periods)
        self.root.update_idletasks()
        
        # Atualizar status final e log
        if success_count == total_periods:
            self.status_var.set(f"✅ {company} [{report_type}]: {success_count}/{total_periods} períodos processados com sucesso")
            self.add_log_entry(f"Resumo {company}: {success_count}/{total_periods} períodos processados com sucesso!", "success")
        elif success_count > 0:
            self.status_var.set(f"⚠️ {company} [{report_type}]: {success_count}/{total_periods} períodos processados com alguns erros")
            self.add_log_entry(f"Resumo {company}: {success_count}/{total_periods} períodos processados com alguns erros", "warning")
        else:
            self.status_var.set(f"❌ {company} [{report_type}]: Falha no processamento")
            self.add_log_entry(f"Resumo {company}: Falha no processamento - 0/{total_periods} períodos processados", "error")
        
        # Adicionar arquivos gerados ao log
        self.add_log_entry(f"Arquivos gerados em: C:\\Users\\contr\\OneDrive\\V.S.Code\\Integração_V7\\Consolidado_Empresas\\{company}", "info")
        self.add_log_entry("=" * 60, "info")  # Linha de separação
        self.root.update_idletasks()
        
        # Aguardar um pouco antes de continuar
        self.root.after(1000, lambda: None)

    def process_ouro_mediano_consolidation(self):
        """Processa a consolidação Ouro Mediano"""
        selected_companies = [self.company_listbox.get(i) for i in self.company_listbox.curselection()]
        selected_years, selected_months = self.get_selected_years_months()
        
        # Se nenhuma empresa selecionada, usar todas as disponíveis
        if not selected_companies:
            if self.ranking_ouro_mediano_processor:
                selected_companies = self.ranking_ouro_mediano_processor.find_available_companies()
            if not selected_companies:
                messagebox.showwarning("Aviso", "Nenhuma empresa encontrada com relatórios Ranking_Por_Empresa.")
                return
        
        # Se nenhum período selecionado, usar todos os disponíveis
        selected_periods = None
        if selected_years and selected_months:
            selected_periods = [f"{m}_{y}" for y in selected_years for m in selected_months]
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        try:
            # Atualizar progresso
            self.update_progress("Iniciando consolidação Ouro Mediano...", 0, 1)
            
            # Processar consolidação
            if self.ranking_ouro_mediano_processor:
                df_consolidated = self.ranking_ouro_mediano_processor.process_consolidation(
                    selected_companies=selected_companies, 
                    selected_periods=selected_periods
                )
                
                if df_consolidated is not None and not df_consolidated.empty:
                    success = self.ranking_ouro_mediano_processor.create_consolidated_report(
                        df_consolidated, 
                        selected_periods=selected_periods
                    )
                    
                    if success:
                        self.update_progress("Consolidação Ouro Mediano concluída com sucesso!", 1, 1)
                        self.status_var.set(f"✅ Consolidação Ouro Mediano concluída: {len(df_consolidated)} registros encontrados")
                        messagebox.showinfo("Sucesso", f"Consolidação Ouro Mediano concluída com sucesso!\n\n{len(df_consolidated)} registros encontrados\n\nArquivo salvo em: Ranking_Ouro_Mediano/")
                    else:
                        self.update_progress("Erro na criação do relatório consolidado", 1, 1)
                        self.status_var.set("❌ Erro na criação do relatório consolidado")
                        messagebox.showerror("Erro", "Erro ao criar o relatório consolidado. Verifique os logs para mais detalhes.")
                else:
                    self.update_progress("Nenhum registro Ouro Mediano encontrado", 1, 1)
                    self.status_var.set("⚠️ Nenhum registro Ouro Mediano encontrado")
                    messagebox.showwarning("Aviso", "Nenhum registro que atenda aos critérios Ouro Mediano foi encontrado.\n\nCritérios:\n- Fase: 'Ouro' ou 'Ouro C'\n- Status: 'Mediano'\n- Ponto acumulado: entre 3.97 e 3.99")
            else:
                self.status_var.set("❌ Processador Ouro Mediano não inicializado")
                messagebox.showerror("Erro", "Processador Ouro Mediano não foi inicializado corretamente.")
                
        except Exception as e:
            logging.error(f"Erro na consolidação Ouro Mediano: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            self.status_var.set("❌ Erro na consolidação Ouro Mediano")
            messagebox.showerror("Erro", f"Erro na consolidação Ouro Mediano:\n{str(e)}")

    def process_km_proporcional(self):
        """Processa o Ranking_Km_Proporcional"""
        selected_companies = [self.company_listbox.get(i) for i in self.company_listbox.curselection()]
        selected_years, selected_months = self.get_selected_years_months()
        
        # Se nenhuma empresa selecionada, usar todas as disponíveis
        if not selected_companies:
            if self.ranking_km_proporcional_processor:
                selected_companies = self.ranking_km_proporcional_processor.find_available_companies()
            if not selected_companies:
                messagebox.showwarning("Aviso", "Nenhuma empresa encontrada com relatórios Ranking_Km_Proporcional.")
                return
        
        # Se nenhum período selecionado, usar todos os disponíveis
        selected_periods = None
        if selected_years and selected_months:
            selected_periods = [f"{m}_{y}" for y in selected_years for m in selected_months]
        
        # Verificar arquivos em uso
        files_in_use = check_excel_files_in_use(self.output_base_dir)
        if files_in_use:
            result = messagebox.askyesno(
                "Arquivos em Uso", 
                f"Encontrados {len(files_in_use)} arquivo(s) Excel em uso.\n\n"
                "Deseja continuar mesmo assim? (O script tentará criar arquivos com nomes alternativos)"
            )
            if not result:
                return
        
        try:
            # Atualizar progresso
            self.update_progress("Iniciando processamento Ranking_Km_Proporcional...", 0, 1)
            
            # Processar Ranking_Km_Proporcional
            if self.ranking_km_proporcional_processor:
                for company in selected_companies:
                    for period in selected_periods:
                        result = self.ranking_km_proporcional_processor.process_company_period(company, period)
                        if result is not None:
                            success = True
                        else:
                            success = False
                        if success:
                            self.update_progress(f"Concluído: {company} - {period} [Ranking_Km_Proporcional]", 1, 1)
                            self.status_var.set(f"✅ Ranking_Km_Proporcional concluído: {period}")
                            messagebox.showinfo("Sucesso", f"Ranking_Km_Proporcional concluído com sucesso!\n\n{period}\n\nArquivo salvo em: Ranking_Km_Proporcional/{company}/")
                        else:
                            self.update_progress("Erro na criação do relatório Ranking_Km_Proporcional", 1, 1)
                            self.status_var.set(f"❌ Erro na criação do relatório Ranking_Km_Proporcional para {company} - {period}")
                            messagebox.showerror("Erro", f"Erro ao criar o relatório Ranking_Km_Proporcional para {company} - {period}. Verifique os logs para mais detalhes.")
            else:
                self.status_var.set("❌ Processador Ranking_Km_Proporcional não inicializado")
                messagebox.showerror("Erro", "Processador Ranking_Km_Proporcional não foi inicializado corretamente.")
                
        except Exception as e:
            logging.error(f"Erro no processamento Ranking_Km_Proporcional: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            self.status_var.set("❌ Erro no processamento Ranking_Km_Proporcional")
            messagebox.showerror("Erro", f"Erro no processamento Ranking_Km_Proporcional:\n{str(e)}")

    def open_rpp_insuficientes_modal(self):
        """Abre o modal para configurar a geração do Relatório de Motoristas Insuficientes"""
        # Criar janela modal
        modal = tk.Toplevel(self.root)
        modal.title("Gerar Relatório de Motoristas Insuficientes")
        modal.geometry("550x280")
        modal.resizable(False, False)
        modal.transient(self.root)  # Modal sempre acima da janela principal
        modal.grab_set()  # Bloqueia interação com a janela principal
        
        # Centralizar o modal
        modal.update_idletasks()
        x = self.root.winfo_x() + (self.root.winfo_width() // 2) - (modal.winfo_width() // 2)
        y = self.root.winfo_y() + (self.root.winfo_height() // 2) - (modal.winfo_height() // 2)
        modal.geometry(f"+{x}+{y}")
        
        # Frame principal do modal
        modal_frame = ttk.Frame(modal, padding=20)
        modal_frame.pack(fill=tk.BOTH, expand=True)
        
        # Título/Descrição
        desc_label = ttk.Label(modal_frame, text="Configure os parâmetros para gerar o relatório consolidado de motoristas insuficientes:", wraplength=500)
        desc_label.grid(row=0, column=0, columnspan=3, pady=(0, 15), sticky="w")
        
        # Campo: Caminho para Ranking_Por_Empresa
        ttk.Label(modal_frame, text="Caminho Ranking_Por_Empresa:").grid(row=1, column=0, sticky="w", pady=5)
        
        caminho_entry = ttk.Entry(modal_frame, width=50)
        # Preencher com o caminho padrão baseado no diretório de saída
        default_ranking_path = os.path.join(self.output_base_dir, "Ranking_Por_Empresa") if self.output_base_dir else ""
        caminho_entry.insert(0, default_ranking_path)
        caminho_entry.grid(row=1, column=1, sticky="ew", padx=5, pady=5)
        
        def browse_ranking_path():
            dir_path = filedialog.askdirectory(
                title="Selecione a pasta Ranking_Por_Empresa",
                initialdir=default_ranking_path if os.path.exists(default_ranking_path) else self.output_base_dir
            )
            if dir_path:
                caminho_entry.delete(0, tk.END)
                caminho_entry.insert(0, dir_path)
        
        ttk.Button(modal_frame, text="Procurar", command=browse_ranking_path).grid(row=1, column=2, padx=5, pady=5)
        
        # Campo: Ano
        ttk.Label(modal_frame, text="Ano:").grid(row=2, column=0, sticky="w", pady=5)
        ano_entry = ttk.Entry(modal_frame, width=20)
        ano_entry.insert(0, str(datetime.now().year))  # Ano atual como padrão
        ano_entry.grid(row=2, column=1, sticky="w", padx=5, pady=5)
        
        # Campo: Mês
        ttk.Label(modal_frame, text="Mês:").grid(row=3, column=0, sticky="w", pady=5)
        meses = ["Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", 
                 "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
        mes_combobox = ttk.Combobox(modal_frame, values=meses, state="readonly", width=17)
        mes_combobox.set(meses[datetime.now().month - 1])  # Mês atual como padrão
        mes_combobox.grid(row=3, column=1, sticky="w", padx=5, pady=5)
        
        # Configurar expansão da coluna 1
        modal_frame.columnconfigure(1, weight=1)
        
        # Frame para botões de ação
        btn_frame = ttk.Frame(modal_frame)
        btn_frame.grid(row=4, column=0, columnspan=3, pady=(20, 0))
        
        def gerar_relatorio():
            caminho = caminho_entry.get().strip()
            ano = ano_entry.get().strip()
            mes = mes_combobox.get()
            
            # Validações
            if not caminho:
                messagebox.showwarning("Aviso", "Por favor, informe o caminho para a pasta Ranking_Por_Empresa.", parent=modal)
                return
            if not os.path.exists(caminho):
                messagebox.showwarning("Aviso", f"O caminho informado não existe:\n{caminho}", parent=modal)
                return
            if not ano or not ano.isdigit():
                messagebox.showwarning("Aviso", "Por favor, informe um ano válido.", parent=modal)
                return
            if not mes:
                messagebox.showwarning("Aviso", "Por favor, selecione um mês.", parent=modal)
                return
            
            # Fechar modal e processar
            modal.destroy()
            self.process_rpp_insuficientes(caminho, ano, mes)
        
        ttk.Button(btn_frame, text="Gerar Relatório", command=gerar_relatorio).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="Cancelar", command=modal.destroy).pack(side=tk.LEFT, padx=10)
        
        # Foco no primeiro campo
        caminho_entry.focus_set()

    def process_rpp_insuficientes(self, ranking_path, ano, mes):
        """Processa e gera o Relatório de Motoristas Insuficientes"""
        try:
            self.add_log_entry("=" * 60, "info")
            self.add_log_entry(f"🚀 Iniciando geração do Relatório de Motoristas Insuficientes", "header")
            self.add_log_entry(f"📁 Caminho: {ranking_path}", "info")
            self.add_log_entry(f"📅 Período: {mes}/{ano}", "info")
            logging.info(f"Iniciando Relatório de Motoristas Insuficientes - Caminho: {ranking_path}, Período: {mes}/{ano}")
            
            self.update_progress("Iniciando processamento de Motoristas Insuficientes...", 0, 1)
            self.root.update_idletasks()
            
            # Verificar se o diretório base existe
            if not os.path.exists(ranking_path):
                msg_erro = f"O diretório base não existe:\n{ranking_path}"
                messagebox.showerror("Erro", msg_erro)
                self.add_log_entry(f"❌ {msg_erro}", "error")
                logging.error(f"Diretório base não existe: {ranking_path}")
                return
            
            # Identificar empresas (subpastas dentro de Ranking_Por_Empresa)
            empresas = []
            try:
                for item in os.listdir(ranking_path):
                    item_path = os.path.join(ranking_path, item)
                    if os.path.isdir(item_path):
                        empresas.append(item)
                logging.info(f"Empresas encontradas no diretório: {empresas}")
            except Exception as e:
                logging.error(f"Erro ao listar diretório {ranking_path}: {str(e)}")
                raise Exception(f"Erro ao listar diretório {ranking_path}: {str(e)}")
            
            if not empresas:
                msg_aviso = f"Nenhuma subpasta de empresa encontrada em:\n{ranking_path}\n\nVerifique se o caminho está correto."
                messagebox.showwarning("Aviso", msg_aviso)
                self.add_log_entry("⚠️ Nenhuma empresa (subpasta) encontrada no diretório.", "warning")
                logging.warning(f"Nenhuma subpasta de empresa encontrada em: {ranking_path}")
                return
            
            self.add_log_entry(f"📊 Empresas encontradas: {len(empresas)} - {', '.join(sorted(empresas))}", "info")
            
            # Dicionário para armazenar DataFrames de cada empresa
            dados_empresas = {}
            empresas_processadas = 0
            empresas_com_erro = []
            empresas_sem_pasta = []
            empresas_sem_arquivo = []
            
            for i, empresa in enumerate(empresas):
                self.update_progress(f"Processando {empresa}...", i, len(empresas))
                self.root.update_idletasks()
                
                # Construir caminho: Empresa/Ano/Mês
                caminho_ano = os.path.join(ranking_path, empresa, ano)
                caminho_mes = os.path.join(ranking_path, empresa, ano, mes)
                
                # Verificar se a pasta do ano existe
                if not os.path.exists(caminho_ano):
                    self.add_log_entry(f"⚠️ Pasta do ano {ano} não encontrada para {empresa}", "warning")
                    logging.warning(f"Pasta do ano não encontrada: {caminho_ano}")
                    empresas_sem_pasta.append(f"{empresa} (pasta {ano} não existe)")
                    continue
                
                # Verificar se a pasta do mês existe
                if not os.path.exists(caminho_mes):
                    self.add_log_entry(f"⚠️ Pasta do mês {mes} não encontrada para {empresa}/{ano}", "warning")
                    logging.warning(f"Pasta do mês não encontrada: {caminho_mes}")
                    empresas_sem_pasta.append(f"{empresa} (pasta {ano}/{mes} não existe)")
                    continue
                
                # Listar todos os arquivos na pasta do mês para debug
                try:
                    arquivos_na_pasta = os.listdir(caminho_mes)
                    logging.info(f"Arquivos encontrados em {caminho_mes}: {arquivos_na_pasta}")
                except Exception as e:
                    self.add_log_entry(f"❌ Erro ao listar pasta {caminho_mes}: {str(e)}", "error")
                    logging.error(f"Erro ao listar pasta {caminho_mes}: {str(e)}")
                    empresas_com_erro.append(f"{empresa} (erro ao listar pasta: {str(e)})")
                    continue
                
                # Procurar arquivo .xlsx na pasta do mês
                arquivo_encontrado = None
                arquivos_xlsx = []
                
                for arquivo in arquivos_na_pasta:
                    if arquivo.endswith('.xlsx') and not arquivo.startswith('~$'):
                        arquivos_xlsx.append(arquivo)
                        # Padrão esperado: Ranking_Por_Empresa_{Empresa}_{Mes}_{Ano}.xlsx
                        # Aceita variações com sufixo de versão (ex: _1.0.xlsx)
                        arquivo_encontrado = os.path.join(caminho_mes, arquivo)
                        break
                
                # Se não encontrou arquivo .xlsx, verificar se há outros arquivos
                if not arquivo_encontrado:
                    if arquivos_na_pasta:
                        self.add_log_entry(f"⚠️ Pasta {empresa}/{ano}/{mes} contém arquivos, mas nenhum .xlsx válido: {arquivos_na_pasta}", "warning")
                        logging.warning(f"Nenhum .xlsx válido em {caminho_mes}. Arquivos encontrados: {arquivos_na_pasta}")
                    else:
                        self.add_log_entry(f"⚠️ Pasta {empresa}/{ano}/{mes} está vazia (nenhum arquivo encontrado)", "warning")
                        logging.warning(f"Pasta vazia: {caminho_mes}")
                    empresas_sem_arquivo.append(f"{empresa} (nenhum .xlsx em {ano}/{mes})")
                    continue
                
                # Ler o arquivo Excel
                try:
                    self.add_log_entry(f"📖 Lendo arquivo: {os.path.basename(arquivo_encontrado)}", "info")
                    logging.info(f"Lendo arquivo Excel: {arquivo_encontrado}")
                    
                    df = pd.read_excel(arquivo_encontrado)
                    
                    if df.empty:
                        self.add_log_entry(f"⚠️ Arquivo vazio para {empresa}: {os.path.basename(arquivo_encontrado)}", "warning")
                        logging.warning(f"Arquivo Excel vazio: {arquivo_encontrado}")
                        empresas_sem_arquivo.append(f"{empresa} (arquivo .xlsx vazio)")
                        continue
                    
                    dados_empresas[empresa] = {
                        'df': df,
                        'arquivo': arquivo_encontrado
                    }
                    empresas_processadas += 1
                    self.add_log_entry(f"✅ {empresa}: {len(df)} registros carregados de {os.path.basename(arquivo_encontrado)}", "success")
                    logging.info(f"Empresa {empresa} carregada com sucesso: {len(df)} registros")
                    
                except Exception as e:
                    self.add_log_entry(f"❌ Erro ao ler arquivo de {empresa}: {str(e)}", "error")
                    logging.error(f"Erro ao ler arquivo {arquivo_encontrado}: {str(e)}")
                    empresas_com_erro.append(f"{empresa} (erro na leitura: {str(e)})")
                    continue
            
            # Verificar se encontrou dados
            if not dados_empresas:
                # Criar mensagem detalhada de erro
                msg_detalhes = []
                if empresas_sem_pasta:
                    msg_detalhes.append(f"• Empresas sem pasta {ano}/{mes}: {len(empresas_sem_pasta)}")
                if empresas_sem_arquivo:
                    msg_detalhes.append(f"• Empresas sem arquivo .xlsx: {len(empresas_sem_arquivo)}")
                if empresas_com_erro:
                    msg_detalhes.append(f"• Empresas com erro: {len(empresas_com_erro)}")
                
                msg_completa = "Nenhum dado encontrado para processar.\n\n"
                if msg_detalhes:
                    msg_completa += "Detalhes:\n" + "\n".join(msg_detalhes)
                msg_completa += "\n\nVerifique o log para mais informações."
                
                messagebox.showwarning("Aviso - Nenhum Dado Encontrado", msg_completa)
                self.add_log_entry("⚠️ Nenhum dado encontrado para processar.", "warning")
                logging.warning(f"Nenhum dado encontrado. Sem pasta: {empresas_sem_pasta}, Sem arquivo: {empresas_sem_arquivo}, Com erro: {empresas_com_erro}")
                self.update_progress("Processamento finalizado sem dados", 1, 1)
                return
            
            # Criar diretório de saída: Saída/RPP_Insuficientes
            # Usar o diretório pai do Ranking_Por_Empresa ou o diretório de saída configurado
            dir_saida = os.path.dirname(ranking_path)  # Diretório pai de Ranking_Por_Empresa
            if not dir_saida or dir_saida == ranking_path:
                dir_saida = self.output_base_dir if self.output_base_dir else os.getcwd()
            
            dir_insuficientes = os.path.join(dir_saida, "RPP_Insuficientes")
            os.makedirs(dir_insuficientes, exist_ok=True)
            
            # Nome do arquivo de saída
            arquivo_saida = os.path.join(dir_insuficientes, "Relatório_Por_Empresa_Insuficientes.xlsx")
            
            self.add_log_entry(f"📁 Criando relatório em: {arquivo_saida}", "info")
            
            # Criar workbook
            wb = Workbook()
            
            # Remover a aba padrão criada automaticamente
            default_sheet = wb.active
            
            # ========== ABA 1: Todas As Empresas ==========
            ws_todas = wb.create_sheet("Todas As Empresas")
            
            # Estilos
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            linha_atual = 1
            primeira_empresa = True
            
            for empresa in sorted(dados_empresas.keys()):
                df = dados_empresas[empresa]['df']
                
                # Se não for a primeira empresa, pular uma linha em branco
                if not primeira_empresa:
                    linha_atual += 1  # Linha em branco
                
                # Escrever cabeçalho
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws_todas.cell(row=linha_atual, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                linha_atual += 1
                
                # Escrever dados
                for row_idx, row in df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_todas.cell(row=linha_atual, column=col_idx, value=value)
                        cell.border = thin_border
                    linha_atual += 1
                
                primeira_empresa = False
            
            # Ajustar largura das colunas na aba "Todas As Empresas"
            for column_cells in ws_todas.columns:
                max_length = 0
                column_letter = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_todas.column_dimensions[column_letter].width = adjusted_width
            
            # ========== ABAS POR EMPRESA ==========
            for empresa in sorted(dados_empresas.keys()):
                df = dados_empresas[empresa]['df']
                
                # Criar nome seguro para a aba (max 31 caracteres, sem caracteres especiais)
                nome_aba = re.sub(r'[\\/*?:\[\]]', '', empresa)[:31]
                
                ws_empresa = wb.create_sheet(nome_aba)
                
                # Escrever cabeçalho
                for col_idx, col_name in enumerate(df.columns, 1):
                    cell = ws_empresa.cell(row=1, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # Escrever dados
                for row_idx, row in df.iterrows():
                    for col_idx, value in enumerate(row, 1):
                        cell = ws_empresa.cell(row=row_idx + 2, column=col_idx, value=value)
                        cell.border = thin_border
                
                # Ajustar largura das colunas
                for column_cells in ws_empresa.columns:
                    max_length = 0
                    column_letter = column_cells[0].column_letter
                    for cell in column_cells:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_empresa.column_dimensions[column_letter].width = adjusted_width
            
            # Remover a aba padrão vazia
            if default_sheet.title == "Sheet":
                wb.remove(default_sheet)
            
            # Salvar o arquivo
            try:
                wb.save(arquivo_saida)
                self.add_log_entry(f"✅ Relatório salvo com sucesso: {arquivo_saida}", "success")
            except PermissionError:
                # Tentar nome alternativo se o arquivo estiver em uso
                arquivo_alternativo = get_alternative_filename(arquivo_saida)
                wb.save(arquivo_alternativo)
                self.add_log_entry(f"✅ Relatório salvo com nome alternativo: {arquivo_alternativo}", "success")
                arquivo_saida = arquivo_alternativo
            
            # Atualizar progresso final
            self.update_progress("Processamento concluído!", 1, 1)
            
            # Calcular totais de problemas
            total_problemas = len(empresas_sem_pasta) + len(empresas_sem_arquivo) + len(empresas_com_erro)
            
            # Resumo
            self.add_log_entry("=" * 60, "info")
            self.add_log_entry(f"📊 RESUMO DO PROCESSAMENTO - Relatório Insuficientes", "header")
            self.add_log_entry(f"📅 Período processado: {mes}/{ano}", "info")
            self.add_log_entry(f"✅ Empresas processadas com sucesso: {empresas_processadas}/{len(empresas)}", "success")
            
            if empresas_sem_pasta:
                self.add_log_entry(f"⚠️ Empresas sem pasta do período ({len(empresas_sem_pasta)}):", "warning")
                for emp in empresas_sem_pasta:
                    self.add_log_entry(f"   - {emp}", "warning")
            
            if empresas_sem_arquivo:
                self.add_log_entry(f"⚠️ Empresas sem arquivo .xlsx ({len(empresas_sem_arquivo)}):", "warning")
                for emp in empresas_sem_arquivo:
                    self.add_log_entry(f"   - {emp}", "warning")
            
            if empresas_com_erro:
                self.add_log_entry(f"❌ Empresas com erro de processamento ({len(empresas_com_erro)}):", "error")
                for err in empresas_com_erro:
                    self.add_log_entry(f"   - {err}", "error")
            
            self.add_log_entry(f"📁 Arquivo gerado: {arquivo_saida}", "info")
            self.add_log_entry("=" * 60, "info")
            
            logging.info(f"Relatório Insuficientes concluído: {empresas_processadas} empresas processadas, {total_problemas} com problemas")
            
            # Mensagem de sucesso
            self.status_var.set(f"✅ Relatório Insuficientes gerado: {empresas_processadas}/{len(empresas)} empresas")
            
            msg_sucesso = f"Relatório gerado com sucesso!\n\n"
            msg_sucesso += f"Total de empresas encontradas: {len(empresas)}\n"
            msg_sucesso += f"Empresas processadas: {empresas_processadas}\n"
            if total_problemas > 0:
                msg_sucesso += f"\n⚠️ Empresas com problemas: {total_problemas}\n"
                if empresas_sem_pasta:
                    msg_sucesso += f"  - Sem pasta {ano}/{mes}: {len(empresas_sem_pasta)}\n"
                if empresas_sem_arquivo:
                    msg_sucesso += f"  - Sem arquivo .xlsx: {len(empresas_sem_arquivo)}\n"
                if empresas_com_erro:
                    msg_sucesso += f"  - Com erro: {len(empresas_com_erro)}\n"
            msg_sucesso += f"\nArquivo salvo em:\n{arquivo_saida}"
            
            messagebox.showinfo("Sucesso", msg_sucesso)
            
        except Exception as e:
            logging.error(f"Erro ao gerar Relatório de Motoristas Insuficientes: {str(e)}")
            logging.error(f"Traceback completo: {traceback.format_exc()}")
            self.add_log_entry(f"❌ Erro crítico: {str(e)}", "error")
            self.status_var.set("❌ Erro ao gerar Relatório de Motoristas Insuficientes")
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{str(e)}")
            self.update_progress("Erro no processamento", 1, 1)

    def generate_pdf_report(self):
        """Gera um relatório PDF com o log de processamento e estatísticas"""
        try:
            # Usa o diretório de saída como padrão, ou o diretório atual
            initial_dir = self.output_base_dir if self.output_base_dir and os.path.exists(self.output_base_dir) else os.getcwd()
            
            # Solicitar local para salvar o PDF
            pdf_path = filedialog.asksaveasfilename(
                title="Salvar Relatório PDF",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                initialdir=initial_dir,
                initialfile=f"Relatorio_Processamento_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            if not pdf_path:
                return
            
            # Criar o documento PDF
            doc = SimpleDocTemplate(pdf_path, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=14,
                spaceAfter=20,
                textColor=colors.darkgreen
            )
            
            header_style = ParagraphStyle(
                'CustomHeader',
                parent=styles['Heading3'],
                fontSize=12,
                spaceAfter=10,
                textColor=colors.darkred
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=6
            )
            
            # Título principal
            story.append(Paragraph("RELATÓRIO DE PROCESSAMENTO", title_style))
            story.append(Spacer(1, 20))
            
            # Informações gerais
            story.append(Paragraph("Informações Gerais", subtitle_style))
            story.append(Paragraph(f"<b>Data e Hora:</b> {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", normal_style))
            story.append(Paragraph(f"<b>Diretório Base:</b> {self.base_dir}", normal_style))
            story.append(Paragraph(f"<b>Diretório de Saída:</b> {self.output_base_dir}", normal_style))
            story.append(Spacer(1, 20))
            
            # Estatísticas do processamento
            story.append(Paragraph("Estatísticas do Processamento", subtitle_style))
            
            # Contar entradas do log
            log_content = self.log_text.get(1.0, tk.END)
            log_lines = log_content.strip().split('\n')
            
            total_entries = len([line for line in log_lines if line.strip()])
            success_entries = len([line for line in log_lines if '✅' in line])
            error_entries = len([line for line in log_lines if '❌' in line])
            warning_entries = len([line for line in log_lines if '⚠️' in line])
            
            stats_data = [
                ['Métrica', 'Quantidade'],
                ['Total de Entradas', str(total_entries)],
                ['Sucessos', str(success_entries)],
                ['Erros', str(error_entries)],
                ['Avisos', str(warning_entries)]
            ]
            
            stats_table = Table(stats_data, colWidths=[2*inch, 1*inch])
            stats_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(stats_table)
            story.append(Spacer(1, 20))
            
            # Log de processamento
            story.append(Paragraph("Log de Processamento", subtitle_style))
            
            # Dividir o log em páginas se necessário
            log_entries = []
            for line in log_lines:
                if line.strip():
                    # Remover caracteres especiais que podem causar problemas no PDF
                    clean_line = line.replace('✅', '[SUCESSO]').replace('❌', '[ERRO]').replace('⚠️', '[AVISO]').replace('🚀', '[INICIO]').replace('⚙️', '[PROCESSANDO]').replace('📊', '[INFO]').replace('ℹ️', '[INFO]')
                    log_entries.append(Paragraph(clean_line, normal_style))
            
            # Adicionar entradas do log em lotes para evitar páginas muito grandes
            batch_size = 50
            for i in range(0, len(log_entries), batch_size):
                batch = log_entries[i:i+batch_size]
                story.extend(batch)
                if i + batch_size < len(log_entries):
                    story.append(PageBreak())
                    story.append(Paragraph("Log de Processamento (continuação)", header_style))
            
            # Resumo final
            story.append(PageBreak())
            story.append(Paragraph("Resumo Final", subtitle_style))
            
            if success_entries > error_entries:
                summary_text = f"✅ Processamento concluído com sucesso! {success_entries} operações bem-sucedidas."
            elif error_entries > 0:
                summary_text = f"⚠️ Processamento concluído com {error_entries} erros. Verifique o log para detalhes."
            else:
                summary_text = "ℹ️ Nenhuma operação foi registrada no log."
            
            story.append(Paragraph(summary_text, normal_style))
            story.append(Spacer(1, 20))
            
            # Gerar o PDF
            doc.build(story)
            
            # Mostrar mensagem de sucesso
            messagebox.showinfo("Sucesso", f"Relatório PDF gerado com sucesso!\nArquivo salvo em: {pdf_path}")
            
            # Adicionar entrada no log
            self.add_log_entry(f"Relatório PDF gerado: {pdf_path}", "success")
            
        except Exception as e:
            error_msg = f"Erro ao gerar PDF: {str(e)}"
            messagebox.showerror("Erro", error_msg)
            self.add_log_entry(error_msg, "error")
            logging.error(f"Erro ao gerar PDF: {traceback.format_exc()}")


if __name__ == '__main__':
    try:
        root = tk.Tk()
        app = UnifiedProcessorGUI(root)
        root.mainloop()
    except tk.TclError as e:
        error_msg = f"""
ERRO: Tcl/Tk não está instalado corretamente!

O tkinter requer os arquivos Tcl/Tk que não foram encontrados na instalação do Python.

SOLUÇÕES:

1. REINSTALAR PYTHON (Recomendado):
   - Baixe o Python 3.14 do site oficial: https://www.python.org/downloads/
   - Durante a instalação, certifique-se de marcar "tcl/tk and IDLE"
   - Ou use o instalador completo que inclui Tcl/Tk

2. INSTALAR TCL/TK SEPARADAMENTE:
   - Baixe ActiveTcl de: https://www.activestate.com/products/tcl/
   - Instale e configure as variáveis de ambiente

3. USAR PYTHON VIA PYTHON.ORG:
   - Desinstale a versão atual
   - Instale a versão do python.org (não Windows Store)

Erro detalhado: {str(e)}
"""
        print(error_msg)
        logging.error(f"Erro ao inicializar tkinter: {str(e)}")
        input("Pressione Enter para sair...")



